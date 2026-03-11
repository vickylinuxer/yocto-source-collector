#!/usr/bin/env python3
"""
source_audit.py — Collect, verify, and report on Yocto image source files.

Usage:
  python3 yocto/source_audit.py <command> -b BUILD -m IMAGE [options]

Commands:
  all       Collect → verify → add missing sources → compile test → report → archive
  collect   Collect sources into ./output/sources/
  verify    DWARF cross-check + coverage check (+ optional compile/audit)
  report    Generate interactive HTML report
  archive   Create per-recipe source distribution tarballs
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import shlex
import shutil
import subprocess
import sys
import tarfile
import tempfile
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


# ═══════════════════════════════════════════════════════════════════════════════
# Constants
# ═══════════════════════════════════════════════════════════════════════════════

SOURCE_EXTS = {".c", ".h", ".S", ".s", ".cpp", ".cc", ".cxx", ".C"}

KERNEL_IMAGE_GLOBS = ("bzImage*", "zImage*", "Image", "vmlinuz*",
                      "uImage*", "fitImage*", "vmlinux")

_ATTR_RE = re.compile(r"DW_AT_(?:name|comp_dir)\s*:\s+(?:\(indirect [^\)]+\):\s*)?(.+)")

_COMPILER_RE = re.compile(
    r'\b(?:[a-z0-9_]+-[a-z0-9_]+-(?:linux-)?(?:gcc|g\+\+|clang|clang\+\+)'
    r'|gcc(?:-\d+(?:\.\d+)*)?|g\+\+(?:-\d+(?:\.\d+)*)?'
    r'|clang(?:-\d+(?:\.\d+)*)?|clang\+\+(?:-\d+(?:\.\d+)*)?|cc|c\+\+)\b'
)
_DEPTH_RE = re.compile(r"make\[(\d+)\]: (Entering|Leaving) directory '(.+)'")
_COMPILED_EXTS = frozenset({".c", ".S", ".s", ".cc", ".cpp", ".cxx"})

TYPE_LABEL = {
    "userspace":     "Userspace",
    "kernel_image":  "Kernel image",
    "kernel_module": "Kernel module",
    "no_source":     "No compiled source",
}

TYPE_COLOR = {
    "userspace":     "#4f86c6",
    "kernel_image":  "#e07b39",
    "kernel_module": "#6dbf67",
    "no_source":     "#b0b0b0",
}

_COPYLEFT_KEYWORDS = {"GPL", "LGPL", "AGPL", "EUPL", "MPL", "CDDL", "OSL"}


def is_copyleft(license_str: str) -> bool:
    upper = license_str.upper()
    return any(kw in upper for kw in _COPYLEFT_KEYWORDS)


# ── License obligation categorization ──

_SPDX_OBLIGATIONS: dict[str, frozenset[str]] = {
    "GPL-2.0-only":        frozenset({"source_distribution", "attribution"}),
    "GPL-2.0-or-later":    frozenset({"source_distribution", "attribution"}),
    "GPL-3.0-only":        frozenset({"source_distribution", "attribution", "patent_grant"}),
    "GPL-3.0-or-later":    frozenset({"source_distribution", "attribution", "patent_grant"}),
    "LGPL-2.0-only":       frozenset({"object_linking", "attribution"}),
    "LGPL-2.0-or-later":   frozenset({"object_linking", "attribution"}),
    "LGPL-2.1-only":       frozenset({"object_linking", "attribution"}),
    "LGPL-2.1-or-later":   frozenset({"object_linking", "attribution"}),
    "LGPL-3.0-only":       frozenset({"object_linking", "attribution", "patent_grant"}),
    "LGPL-3.0-or-later":   frozenset({"object_linking", "attribution", "patent_grant"}),
    "AGPL-3.0-only":       frozenset({"source_distribution", "network_copyleft", "attribution", "patent_grant"}),
    "AGPL-3.0-or-later":   frozenset({"source_distribution", "network_copyleft", "attribution", "patent_grant"}),
    "MPL-2.0":             frozenset({"source_distribution", "attribution", "patent_grant"}),
    "EPL-1.0":             frozenset({"source_distribution", "attribution", "patent_grant"}),
    "EPL-2.0":             frozenset({"source_distribution", "attribution", "patent_grant"}),
    "CDDL-1.0":            frozenset({"source_distribution", "attribution", "patent_grant"}),
    "CDDL-1.1":            frozenset({"source_distribution", "attribution", "patent_grant"}),
    "EUPL-1.1":            frozenset({"source_distribution", "network_copyleft", "attribution"}),
    "EUPL-1.2":            frozenset({"source_distribution", "network_copyleft", "attribution"}),
    "OSL-3.0":             frozenset({"source_distribution", "network_copyleft", "attribution", "patent_grant"}),
    "Apache-2.0":          frozenset({"attribution", "patent_grant", "permissive"}),
    "MIT":                 frozenset({"attribution", "permissive"}),
    "BSD-2-Clause":        frozenset({"attribution", "permissive"}),
    "BSD-3-Clause":        frozenset({"attribution", "permissive"}),
    "ISC":                 frozenset({"attribution", "permissive"}),
    "Zlib":                frozenset({"attribution", "permissive"}),
    "BSL-1.0":             frozenset({"permissive"}),
    "Unlicense":           frozenset({"permissive"}),
    "CC0-1.0":             frozenset({"permissive"}),
}

_LEGACY_LICENSE_MAP: dict[str, str] = {
    "GPLv2":        "GPL-2.0-only",
    "GPLv2+":       "GPL-2.0-or-later",
    "GPLv2.0":      "GPL-2.0-only",
    "GPLv2.0+":     "GPL-2.0-or-later",
    "GPLv3":        "GPL-3.0-only",
    "GPLv3+":       "GPL-3.0-or-later",
    "LGPLv2":       "LGPL-2.0-only",
    "LGPLv2+":      "LGPL-2.0-or-later",
    "LGPLv2.1":     "LGPL-2.1-only",
    "LGPLv2.1+":    "LGPL-2.1-or-later",
    "LGPLv3":       "LGPL-3.0-only",
    "LGPLv3+":      "LGPL-3.0-or-later",
    "AGPLv3":       "AGPL-3.0-only",
    "AGPLv3+":      "AGPL-3.0-or-later",
    "MPL-2":        "MPL-2.0",
    "MPLv2":        "MPL-2.0",
    "Apache-2":     "Apache-2.0",
    "Apachev2":     "Apache-2.0",
    "BSD":          "BSD-3-Clause",
    "BSD-2":        "BSD-2-Clause",
    "BSD-3":        "BSD-3-Clause",
    "PD":           "Unlicense",
}

# Case-insensitive lookup for SPDX obligations
_SPDX_OBLIGATIONS_LOWER: dict[str, str] = {
    k.lower(): k for k in _SPDX_OBLIGATIONS
}


def _normalize_license_id(lic_id: str) -> str:
    """Normalize a license identifier to its SPDX form."""
    stripped = lic_id.strip()
    if stripped in _SPDX_OBLIGATIONS:
        return stripped
    if stripped in _LEGACY_LICENSE_MAP:
        return _LEGACY_LICENSE_MAP[stripped]
    lower = stripped.lower()
    if lower in _SPDX_OBLIGATIONS_LOWER:
        return _SPDX_OBLIGATIONS_LOWER[lower]
    return stripped


def _heuristic_obligations(lic_id: str) -> frozenset[str]:
    """Keyword-based fallback for unknown license IDs."""
    upper = lic_id.upper()
    obs: set[str] = set()
    if "AGPL" in upper:
        obs.update({"source_distribution", "network_copyleft", "attribution"})
    elif "LGPL" in upper:
        obs.update({"object_linking", "attribution"})
    elif "GPL" in upper:
        obs.update({"source_distribution", "attribution"})
    elif "MPL" in upper or "CDDL" in upper or "EPL" in upper or "OSL" in upper or "EUPL" in upper:
        obs.update({"source_distribution", "attribution"})
    elif "APACHE" in upper:
        obs.update({"attribution", "patent_grant", "permissive"})
    elif any(kw in upper for kw in ("MIT", "BSD", "ISC", "ZLIB")):
        obs.update({"attribution", "permissive"})
    elif any(kw in upper for kw in ("UNLICENSE", "CC0", "PD", "PUBLIC DOMAIN")):
        obs.add("permissive")
    else:
        obs.add("attribution")
    return frozenset(obs)


def classify_license(license_str: str) -> dict:
    """Classify a Yocto license string into obligation categories.

    Splits compound `&`/`|` expressions and returns aggregated obligations.
    """
    all_obligations: set[str] = set()
    licenses: list[dict] = []

    tokens = re.split(r'\s*[&|]\s*', license_str)
    for tok in tokens:
        tok = tok.strip()
        if not tok:
            continue
        normalized = _normalize_license_id(tok)
        if normalized in _SPDX_OBLIGATIONS:
            obs = _SPDX_OBLIGATIONS[normalized]
        else:
            obs = _heuristic_obligations(normalized)
        all_obligations.update(obs)
        licenses.append({"id": normalized, "obligations": sorted(obs)})

    return {
        "obligations": sorted(all_obligations),
        "has_source_distribution": "source_distribution" in all_obligations,
        "has_object_linking": "object_linking" in all_obligations,
        "has_attribution": "attribution" in all_obligations,
        "has_patent_grant": "patent_grant" in all_obligations,
        "has_network_copyleft": "network_copyleft" in all_obligations,
        "is_permissive_only": all_obligations <= {"attribution", "patent_grant", "permissive"},
        "licenses": licenses,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# Data classes
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class KernelInfo:
    build_dir: Path
    src_dir: Path

    @property
    def modules_order(self) -> Path:
        return self.build_dir / "modules.order"

    def module_objs(self) -> set[Path]:
        owned: set[Path] = set()
        if not self.modules_order.exists():
            return owned
        for line in self.modules_order.read_text().splitlines():
            line = line.strip()
            if not line:
                continue
            ko = Path(line)
            mod_file = self.build_dir / ko.parent / (ko.stem + ".mod")
            if mod_file.exists():
                for obj_rel in mod_file.read_text().split():
                    if obj_rel:
                        owned.add(self.build_dir / obj_rel)
        return owned


@dataclass
class PackageInfo:
    installed_name: str
    yocto_pkg: str
    recipe: str
    ver: str
    work_ver: Path | None
    pkg_type: str
    recipe_prefix: str = ""
    kernel: KernelInfo | None = None
    kernel_mod_obj_rels: list[str] = field(default_factory=list)


@dataclass
class CompileCmd:
    cwd:  Path
    cmd:  str
    src:  Path
    obj:  Path | None


@dataclass
class AuditFinding:
    severity: str   # "FAIL" | "WARN" | "INFO"
    code:     str   # e.g. "DWARF_TIMEOUT", "MISSING_FROM_COLLECTION"
    detail:   str


@dataclass
class AuditResult:
    pkg_name:  str
    pkg_type:  str
    verdict:   str  # "PASS" | "WARN" | "FAIL"
    findings:  list[AuditFinding] = field(default_factory=list)
    evidence:  dict = field(default_factory=dict)


# ═══════════════════════════════════════════════════════════════════════════════
# Discovery — auto-detection, pkgdata, manifest, work-dir helpers
# ═══════════════════════════════════════════════════════════════════════════════

# ── bitbake -e helpers ───────────────────────────────────────────────────────

_BB_GLOBAL_VARS = {
    "TMPDIR", "MACHINE", "STAMPS_DIR", "STAGING_BINDIR_NATIVE",
}
_BB_RECIPE_VARS = {"S", "B", "WORKDIR", "STAGING_KERNEL_DIR", "SRC_URI"}


def _parse_bitbake_env(output: str, variables: set[str]) -> dict[str, str]:
    """Extract variable assignments from ``bitbake -e`` output.

    Lines look like: TMPDIR="/path/to/tmp"
    Multi-line values (e.g. SRC_URI) use backslash continuation and end
    with a closing ``"``.  Skips commented lines (prefixed with ``#``).
    """
    result: dict[str, str] = {}
    lines = output.splitlines()
    i = 0
    while i < len(lines):
        line = lines[i]
        if line.startswith("#"):
            i += 1
            continue
        for var in variables:
            prefix = f'{var}="'
            if line.startswith(prefix):
                if line.endswith('"'):
                    # Single-line value
                    result[var] = line[len(prefix):-1]
                else:
                    # Multi-line value — collect until closing "
                    parts = [line[len(prefix):]]
                    i += 1
                    while i < len(lines):
                        cont = lines[i]
                        if cont.endswith('"'):
                            parts.append(cont[:-1])
                            break
                        parts.append(cont)
                        i += 1
                    result[var] = " ".join(p.rstrip(" \\") for p in parts)
                break
        i += 1
    return result


def _is_bitbake_available(build_dir: Path) -> bool:
    """Return True if bitbake is on PATH and the build environment is sourced."""
    return _check_bitbake_env(build_dir) is None


def _bb_cache_dir() -> Path:
    """Return the bitbake -e disk cache directory."""
    return Path("./output/.bb_cache").resolve()


def _bb_cache_load(name: str) -> dict[str, str] | None:
    """Load a cached bitbake -e result from disk. Returns None on miss."""
    path = _bb_cache_dir() / f"{name}.json"
    if path.exists():
        try:
            data = json.loads(path.read_text())
            if isinstance(data, dict):
                return data
        except (json.JSONDecodeError, OSError):
            pass
    return None


def _bb_cache_save(name: str, data: dict[str, str]) -> None:
    """Save a bitbake -e result to disk cache."""
    cache_dir = _bb_cache_dir()
    cache_dir.mkdir(parents=True, exist_ok=True)
    (cache_dir / f"{name}.json").write_text(json.dumps(data))


def _query_bitbake_global(build_dir: Path) -> dict[str, str] | None:
    """Run ``bitbake -e`` and return global variable dict, or None if unavailable.

    Results are cached to disk (``output/.bb_cache/_global.json``) so
    subsequent commands can skip the expensive ``bitbake -e`` call.
    """
    cached = _bb_cache_load("_global")
    if cached is not None:
        return cached
    if not _is_bitbake_available(build_dir):
        return None
    try:
        r = subprocess.run(
            ["bitbake", "-e"],
            capture_output=True, text=True,
            cwd=str(build_dir), timeout=120,
        )
        if r.returncode != 0:
            return None
        result = _parse_bitbake_env(r.stdout, _BB_GLOBAL_VARS)
        if result:
            _bb_cache_save("_global", result)
        return result
    except (subprocess.TimeoutExpired, FileNotFoundError, OSError):
        return None


def _query_bitbake_recipe(recipe: str, build_dir: Path) -> dict[str, str] | None:
    """Run ``bitbake -e <recipe>`` and return recipe variable dict, or None.

    Results are cached to disk (``output/.bb_cache/{recipe}.json``) so
    subsequent commands can skip the expensive per-recipe query.
    """
    cached = _bb_cache_load(recipe)
    if cached is not None:
        return cached
    if not _is_bitbake_available(build_dir):
        return None
    try:
        r = subprocess.run(
            ["bitbake", "-e", recipe],
            capture_output=True, text=True,
            cwd=str(build_dir), timeout=120,
        )
        if r.returncode != 0:
            return None
        result = _parse_bitbake_env(r.stdout, _BB_RECIPE_VARS)
        if result:
            _bb_cache_save(recipe, result)
        return result
    except (subprocess.TimeoutExpired, FileNotFoundError, OSError):
        return None


def _resolve_tmpdir(build_dir: Path,
                    bb_env: dict[str, str] | None = None) -> Path:
    """Auto-detect the Yocto TMPDIR under build_dir.

    Tries ``bitbake -e`` result first (if *bb_env* provided), then checks
    conf/local.conf for TMPDIR assignment, then falls back to scanning for
    tmp*/pkgdata directories, and finally defaults to 'tmp'.
    """
    # Strategy 0: use bitbake -e result
    if bb_env and "TMPDIR" in bb_env:
        p = Path(bb_env["TMPDIR"])
        if p.is_dir():
            return p

    # Strategy 1: parse TMPDIR from conf/local.conf
    local_conf = build_dir / "conf" / "local.conf"
    if local_conf.exists():
        try:
            for line in local_conf.read_text(errors="replace").splitlines():
                line = line.strip()
                if line.startswith("#"):
                    continue
                m = re.match(r'TMPDIR\s*[?:]?=\s*"(.+?)"', line)
                if m:
                    raw = m.group(1)
                    # Expand ${TOPDIR} — the standard Yocto variable
                    expanded = raw.replace("${TOPDIR}", str(build_dir))
                    p = Path(expanded)
                    if not p.is_absolute():
                        p = build_dir / p
                    if p.is_dir():
                        return p
        except OSError:
            pass

    # Strategy 2: scan for tmp*/pkgdata directories
    try:
        for child in sorted(build_dir.iterdir()):
            if (child.is_dir()
                    and child.name.startswith("tmp")
                    and (child / "pkgdata").is_dir()):
                return child
    except OSError:
        pass

    # Strategy 3: default
    return build_dir / "tmp"


def auto_detect_machine(build_dir: Path, tmpdir: Path | None = None) -> str:
    if tmpdir is None:
        tmpdir = _resolve_tmpdir(build_dir)
    pkgdata = tmpdir / "pkgdata"
    if not pkgdata.exists():
        raise SystemExit(f"pkgdata not found: {pkgdata}\n"
                         "Make sure --build-dir points to a Yocto build directory.")
    dirs = [d.name for d in pkgdata.iterdir()
            if d.is_dir() and d.name not in ("sdk", "world")]
    if len(dirs) == 1:
        return dirs[0]
    if len(dirs) == 0:
        raise SystemExit(f"No machine directories found in {pkgdata}")
    raise SystemExit(
        f"Multiple machines found in {pkgdata}: {dirs}\n"
        "Use --machine to specify which one."
    )


def find_manifest(build_dir: Path, image_or_path: str,
                   tmpdir: Path | None = None) -> Path:
    p = Path(image_or_path)
    if p.exists():
        return p.resolve()
    if tmpdir is None:
        tmpdir = _resolve_tmpdir(build_dir)
    deploy = tmpdir / "deploy" / "images"
    if not deploy.exists():
        raise SystemExit(f"Deploy images dir not found: {deploy}")
    candidates: list[Path] = []
    for machine_dir in deploy.iterdir():
        if not machine_dir.is_dir():
            continue
        for m in machine_dir.glob(f"{image_or_path}*.manifest"):
            candidates.append(m)
    if not candidates:
        raise SystemExit(f"No manifest found for '{image_or_path}' under {deploy}")
    no_ts = [c for c in candidates if not re.search(r"\d{14}", c.name)]
    if no_ts:
        return sorted(no_ts)[0]
    return sorted(candidates)[-1]


def build_pkgname_map(pkgdata_runtime: Path) -> dict[str, str]:
    mapping: dict[str, str] = {}
    if not pkgdata_runtime.exists():
        return mapping
    for f in pkgdata_runtime.iterdir():
        if f.suffix == ".packaged" or not f.is_file():
            continue
        try:
            for line in f.read_text(errors="replace").splitlines():
                if line.startswith("PKG_"):
                    key, _, installed_name = line.partition(":")
                    if not installed_name:
                        continue
                    yocto_pkg = key[4:]
                    installed_name = installed_name.strip()
                    if installed_name and installed_name != yocto_pkg:
                        mapping[installed_name] = yocto_pkg
        except OSError:
            pass
    return mapping


def parse_pkgdata_file(path: Path) -> dict[str, str]:
    data: dict[str, str] = {}
    try:
        for line in path.read_text(errors="replace").splitlines():
            k, _, v = line.partition(":")
            k = k.strip()
            if k and " " not in k:
                data[k] = v.strip()
    except OSError:
        pass
    return data


def parse_manifest(manifest_path: Path) -> list[tuple[str, str, str]]:
    result: list[tuple[str, str, str]] = []
    for line in manifest_path.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        parts = line.split()
        pkg  = parts[0]
        arch = parts[1] if len(parts) > 1 else ""
        ver  = parts[2] if len(parts) > 2 else ""
        result.append((pkg, arch, ver))
    return result


def find_work_ver_dir(work: Path, recipe: str, ver: str) -> Path | None:
    candidates = [
        d for d in work.glob(f"*/{recipe}/{ver}")
        if d.parent.parent.parent == work and d.is_dir()
    ]
    if not candidates:
        return None
    for c in candidates:
        if (c / "debugsources.list").exists():
            return c
    for c in candidates:
        if (c / "packages-split").exists():
            return c
    return candidates[0]


def find_kernel_build_dir(work_ver: Path) -> Path | None:
    # Pattern 1: *-build (e.g. linux-raspberrypi-5.15.92-build, linux-yocto-6.1-build)
    for d in work_ver.iterdir():
        if d.is_dir() and d.name.endswith("-build"):
            return d
    # Pattern 2: plain build/ subdirectory (e.g. linux.euto-v9 uses work_ver/build/)
    build_dir = work_ver / "build"
    if build_dir.is_dir():
        return build_dir
    return None


def find_kernel_src_dir(build_dir: Path, machine: str,
                        tmpdir: Path | None = None) -> Path | None:
    if tmpdir is None:
        tmpdir = _resolve_tmpdir(build_dir)
    p = tmpdir / "work-shared" / machine / "kernel-source"
    return p if p.exists() else None


# ── ELF helpers ──────────────────────────────────────────────────────────────

def is_elf(path: Path) -> bool:
    try:
        with open(path, "rb") as f:
            return f.read(4) == b"\x7fELF"
    except OSError:
        return False


def find_installed_elfs(pkg_split_dir: Path) -> list[Path]:
    if not pkg_split_dir.exists():
        return []
    return [
        p for p in pkg_split_dir.rglob("*")
        if p.is_file() and not p.is_symlink() and is_elf(p)
    ]


def find_debug_counterpart(elf: Path, pkg_split_dir: Path) -> Path | None:
    split_root = pkg_split_dir.parent
    rel = elf.relative_to(pkg_split_dir)
    name = elf.name
    canonical = (split_root / (pkg_split_dir.name + "-dbg")
                 / rel.parent / ".debug" / name)
    if canonical.exists():
        return canonical
    for dbg_dir in split_root.iterdir():
        if not (dbg_dir.name.endswith("-dbg") and dbg_dir.is_dir()):
            continue
        for candidate in dbg_dir.rglob(f".debug/{name}"):
            if candidate.is_file():
                return candidate
    return None


# ── DWARF source path extraction ────────────────────────────────────────────

def _extract_dwarf_safe(elf_path: Path, timeout: int = 180) -> tuple[set[str], bool]:
    """Extract DWARF CU source paths from an ELF binary.

    Returns (sources, timed_out).
    """
    try:
        r = subprocess.run(
            ["readelf", "--debug-dump=info", str(elf_path)],
            capture_output=True, text=True, timeout=timeout,
        )
    except subprocess.TimeoutExpired:
        return set(), True
    except FileNotFoundError:
        return set(), False

    sources: set[str] = set()
    cu_name: str | None = None
    cu_comp_dir: str | None = None
    in_cu = False

    def _commit():
        nonlocal cu_name, cu_comp_dir
        if cu_name and cu_comp_dir:
            if cu_name.startswith("/"):
                sources.add(os.path.normpath(cu_name))
            else:
                sources.add(os.path.normpath(
                    cu_comp_dir.rstrip("/") + "/" + cu_name))
        cu_name = cu_comp_dir = None

    for line in r.stdout.splitlines():
        if "DW_TAG_compile_unit" in line:
            _commit()
            in_cu = True
        elif not in_cu:
            continue
        elif "DW_AT_name" in line and cu_name is None:
            m = _ATTR_RE.search(line)
            if m:
                val = m.group(1).strip().rstrip(")")
                if val.endswith((".c", ".h", ".S", ".s", ".cpp", ".cc", ".cxx", ".C")):
                    cu_name = val
        elif "DW_AT_comp_dir" in line and cu_comp_dir is None:
            m = _ATTR_RE.search(line)
            if m:
                cu_comp_dir = m.group(1).strip().rstrip(")")
        if cu_name and cu_comp_dir:
            _commit()
            in_cu = False

    _commit()
    return sources, False


def extract_dwarf_cu_sources(elf_path: Path, timeout: int = 180) -> set[str]:
    """Convenience wrapper: returns just the source set (no timeout flag)."""
    sources, _ = _extract_dwarf_safe(elf_path, timeout)
    return sources


# ── Package classification ──────────────────────────────────────────────────

def is_kernel_recipe(recipe: str) -> bool:
    return recipe.startswith("linux-") or recipe == "linux"


def _classify_kernel_pkg(yocto_pkg: str, work_ver: Path) -> tuple[str, list[str]]:
    pkg_split = work_ver / "packages-split" / yocto_pkg
    if not pkg_split.exists():
        return "no_source", []
    for glob in KERNEL_IMAGE_GLOBS:
        if list(pkg_split.rglob(glob)):
            return "kernel_image", []
    ko_files = list(pkg_split.rglob("*.ko"))
    if ko_files:
        return "kernel_module", [str(f.relative_to(pkg_split)) for f in ko_files]
    return "no_source", []


def _get_module_obj_rels(
    pkg_split: Path, kbuild: Path, ko_rels_in_split: list[str]
) -> list[str]:
    obj_rels: list[str] = []
    for ko_rel in ko_rels_in_split:
        ko_name = Path(ko_rel).name
        for ko_in_build in kbuild.rglob(ko_name):
            if ko_in_build.suffix != ".ko":
                continue
            mod_file = ko_in_build.parent / (ko_in_build.stem + ".mod")
            if mod_file.exists():
                for tok in mod_file.read_text().split():
                    if tok:
                        obj_rels.append(tok)
            break
    return obj_rels


# ── Main discovery entry point ──────────────────────────────────────────────

def discover_packages(
    manifest_path: Path,
    build_dir: Path,
    machine: str,
    verbose: bool = False,
    tmpdir: Path | None = None,
    session: "YoctoSession | None" = None,
) -> list[PackageInfo]:
    if tmpdir is None:
        tmpdir = _resolve_tmpdir(build_dir)
    pkgdata_runtime = tmpdir / "pkgdata" / machine / "runtime"
    work = tmpdir / "work"

    pkgname_map = build_pkgname_map(pkgdata_runtime)
    manifest_pkgs = parse_manifest(manifest_path)

    _work_cache:   dict[tuple[str, str], Path | None] = {}
    _kernel_cache: dict[Path, KernelInfo | None] = {}

    packages: list[PackageInfo] = []

    for installed_pkg, _arch, _ver in manifest_pkgs:
        yocto_pkg = pkgname_map.get(installed_pkg, installed_pkg)

        pkgdata_file = pkgdata_runtime / yocto_pkg
        if not pkgdata_file.exists():
            pkgdata_file = pkgdata_runtime / installed_pkg
        if not pkgdata_file.exists():
            if verbose:
                print(f"  [WARN] no pkgdata for {installed_pkg} (yocto_pkg={yocto_pkg})")
            packages.append(PackageInfo(
                installed_name=installed_pkg, yocto_pkg=yocto_pkg,
                recipe=installed_pkg, ver="unknown",
                work_ver=None, pkg_type="no_source",
            ))
            continue

        data = parse_pkgdata_file(pkgdata_file)
        recipe = data.get("PN", yocto_pkg)
        pv     = data.get("PV", "")
        pr     = data.get("PR", "r0")
        pe     = data.get("PE", "")
        if pv:
            ver = f"{pe}_{pv}-{pr}" if pe else f"{pv}-{pr}"
        else:
            ver = "unknown"

        cache_key = (recipe, ver)
        if cache_key not in _work_cache:
            # Try bitbake WORKDIR first, fallback to heuristic
            bb_workdir: Path | None = None
            if session:
                renv = session.query_recipe_env(recipe)
                bb_wd = renv.get("WORKDIR")
                if bb_wd:
                    p = Path(bb_wd)
                    if p.is_dir():
                        bb_workdir = p
            _work_cache[cache_key] = bb_workdir or find_work_ver_dir(work, recipe, ver)
        work_ver = _work_cache[cache_key]

        if work_ver is None:
            if verbose:
                print(f"  [WARN] work dir not found: {recipe}/{ver}")
            packages.append(PackageInfo(
                installed_name=installed_pkg, yocto_pkg=yocto_pkg,
                recipe=recipe, ver=ver,
                work_ver=None, pkg_type="no_source",
            ))
            continue

        recipe_prefix = f"/usr/src/debug/{recipe}/{ver}/"

        if is_kernel_recipe(recipe):
            pkg_type, ko_rels = _classify_kernel_pkg(yocto_pkg, work_ver)

            kernel: KernelInfo | None = None
            mod_obj_rels: list[str] = []

            if pkg_type in ("kernel_image", "kernel_module"):
                if work_ver not in _kernel_cache:
                    # Try bitbake B for kernel build dir
                    bb_kbuild: Path | None = None
                    bb_ksrc: Path | None = None
                    if session:
                        renv = session.query_recipe_env(recipe)
                        bb_b = renv.get("B")
                        if bb_b:
                            p = Path(bb_b)
                            if p.is_dir():
                                bb_kbuild = p
                        bb_skd = renv.get("STAGING_KERNEL_DIR")
                        if bb_skd:
                            p = Path(bb_skd)
                            if p.is_dir():
                                bb_ksrc = p
                    kbuild = bb_kbuild or find_kernel_build_dir(work_ver)
                    ksrc   = bb_ksrc or find_kernel_src_dir(build_dir, machine, tmpdir)
                    _kernel_cache[work_ver] = (
                        KernelInfo(build_dir=kbuild, src_dir=ksrc)
                        if kbuild and ksrc else None
                    )
                kernel = _kernel_cache[work_ver]

                if kernel is None and verbose:
                    kbuild = find_kernel_build_dir(work_ver)
                    ksrc = find_kernel_src_dir(build_dir, machine, tmpdir)
                    reasons = []
                    if not kbuild:
                        reasons.append(f"no *-build dir in {work_ver}")
                    if not ksrc:
                        reasons.append(f"kernel-source not found at "
                                       f"{tmpdir}/work-shared/{machine}/kernel-source")
                    print(f"  [WARN] {installed_pkg}: kernel={pkg_type} but KernelInfo unavailable"
                          f": {'; '.join(reasons)}")

                if pkg_type == "kernel_module" and kernel:
                    pkg_split = work_ver / "packages-split" / yocto_pkg
                    mod_obj_rels = _get_module_obj_rels(
                        pkg_split, kernel.build_dir, ko_rels)

            packages.append(PackageInfo(
                installed_name=installed_pkg, yocto_pkg=yocto_pkg,
                recipe=recipe, ver=ver, work_ver=work_ver,
                pkg_type=pkg_type, recipe_prefix=recipe_prefix,
                kernel=kernel, kernel_mod_obj_rels=mod_obj_rels,
            ))
            continue

        debugsources = work_ver / "debugsources.list"
        compile_log = work_ver / "temp" / "log.do_compile"
        if not debugsources.exists():
            pkg_type = "no_source"
        elif not compile_log.exists():
            pkg_type = "no_source"
        else:
            pkg_split = work_ver / "packages-split" / yocto_pkg
            pkg_type = "userspace" if find_installed_elfs(pkg_split) else "no_source"

        packages.append(PackageInfo(
            installed_name=installed_pkg, yocto_pkg=yocto_pkg,
            recipe=recipe, ver=ver, work_ver=work_ver,
            pkg_type=pkg_type, recipe_prefix=recipe_prefix,
        ))

    # Inject synthetic kernel-image-image if not already present
    has_kernel_image = any(p.pkg_type == "kernel_image" for p in packages)
    if not has_kernel_image:
        ki_pkgdata = pkgdata_runtime / "kernel-image-image"
        if ki_pkgdata.exists():
            data = parse_pkgdata_file(ki_pkgdata)
            ki_recipe = data.get("PN", "")
            if not ki_recipe:
                if verbose:
                    print(f"  [WARN] Synthetic kernel-image-image: PN missing in {ki_pkgdata}")
            else:
                pv = data.get("PV", "")
                pr = data.get("PR", "r0")
                pe = data.get("PE", "")
                ki_ver = f"{pe}_{pv}-{pr}" if pe else f"{pv}-{pr}" if pv else "unknown"
                cache_key = (ki_recipe, ki_ver)
                if cache_key not in _work_cache:
                    bb_workdir_ki: Path | None = None
                    if session:
                        renv = session.query_recipe_env(ki_recipe)
                        bb_wd = renv.get("WORKDIR")
                        if bb_wd:
                            p = Path(bb_wd)
                            if p.is_dir():
                                bb_workdir_ki = p
                    _work_cache[cache_key] = bb_workdir_ki or find_work_ver_dir(work, ki_recipe, ki_ver)
                ki_work_ver = _work_cache[cache_key]
                if ki_work_ver:
                    if ki_work_ver not in _kernel_cache:
                        bb_kbuild_ki: Path | None = None
                        bb_ksrc_ki: Path | None = None
                        if session:
                            renv = session.query_recipe_env(ki_recipe)
                            bb_b = renv.get("B")
                            if bb_b:
                                p = Path(bb_b)
                                if p.is_dir():
                                    bb_kbuild_ki = p
                            bb_skd = renv.get("STAGING_KERNEL_DIR")
                            if bb_skd:
                                p = Path(bb_skd)
                                if p.is_dir():
                                    bb_ksrc_ki = p
                        kbuild = bb_kbuild_ki or find_kernel_build_dir(ki_work_ver)
                        ksrc = bb_ksrc_ki or find_kernel_src_dir(build_dir, machine, tmpdir)
                        _kernel_cache[ki_work_ver] = (
                            KernelInfo(build_dir=kbuild, src_dir=ksrc)
                            if kbuild and ksrc else None
                        )
                    ki_kernel = _kernel_cache[ki_work_ver]
                    if ki_kernel:
                        ki_prefix = f"/usr/src/debug/{ki_recipe}/{ki_ver}/"
                        packages.append(PackageInfo(
                            installed_name="kernel-image-image",
                            yocto_pkg="kernel-image-image",
                            recipe=ki_recipe, ver=ki_ver,
                            work_ver=ki_work_ver,
                            pkg_type="kernel_image",
                            recipe_prefix=ki_prefix,
                            kernel=ki_kernel,
                        ))
                        if verbose:
                            print(f"  [INFO] Injected synthetic kernel-image-image "
                                  f"(recipe={ki_recipe}, ver={ki_ver})")
                    elif verbose:
                        kbuild = find_kernel_build_dir(ki_work_ver)
                        ksrc = find_kernel_src_dir(build_dir, machine, tmpdir)
                        reasons = []
                        if not kbuild:
                            reasons.append(f"no *-build dir in {ki_work_ver}")
                        if not ksrc:
                            reasons.append(f"kernel-source not found at "
                                           f"{tmpdir}/work-shared/{machine}/kernel-source")
                        print(f"  [WARN] Synthetic kernel-image-image: KernelInfo creation failed"
                              f": {'; '.join(reasons)}")
                elif verbose:
                    print(f"  [WARN] Synthetic kernel-image-image: work_ver not found "
                          f"for {ki_recipe}/{ki_ver}")
        elif verbose:
            print(f"  [WARN] Synthetic kernel-image-image: pkgdata not found "
                  f"at {ki_pkgdata}")

    return packages


# ── Source file utilities ───────────────────────────────────────────────────

def read_debugsources(path: Path) -> list[str]:
    data = path.read_bytes()
    sep  = b"\x00" if b"\x00" in data else b"\n"
    return [
        p.decode("utf-8", errors="replace").strip()
        for p in data.split(sep) if p.strip()
    ]


def list_collected_files(output_dir: Path, pkg: str) -> set[str]:
    d = output_dir / pkg
    if not d.exists():
        return set()
    return {
        str(p.relative_to(d))
        for p in d.rglob("*")
        if p.is_file() and p.suffix in SOURCE_EXTS
    }


def strip_src_root(rel: str) -> str:
    parts = Path(rel).parts
    return str(Path(*parts[1:])) if len(parts) > 1 else rel


def is_build_dir(dirname: str) -> bool:
    n = dirname.lower()
    if n in ("build", "builds", "_build", ".build"):
        return True
    if n.startswith("build-") or n.endswith("-build"):
        return True
    if re.search(r"^[a-z0-9_]+-[a-z0-9_]+-linux", n):
        return True
    return False


def get_installed_files(pkgdata_runtime: Path, yocto_pkg: str,
                        pkg_split: Path | None = None) -> list[dict]:
    pkgdata_file = pkgdata_runtime / yocto_pkg
    if not pkgdata_file.exists():
        return []
    data = parse_pkgdata_file(pkgdata_file)
    raw = data.get("FILES_INFO", "")
    if not raw:
        return []
    try:
        info = json.loads(raw)
    except (json.JSONDecodeError, ValueError):
        return []
    result: list[dict] = []
    for fpath, size in info.items():
        elf = False
        if pkg_split:
            on_disk = pkg_split / fpath.lstrip("/")
            if on_disk.exists() and not on_disk.is_symlink():
                elf = is_elf(on_disk)
        result.append({"path": fpath, "size": size, "is_elf": elf})
    return result


_LICENSE_SKIP_EXTS = {
    ".c", ".h", ".cpp", ".cc", ".cxx", ".py", ".pl", ".rb", ".sh",
    ".m4", ".S", ".s", ".java", ".go", ".rs",
}


def get_license_files(build_dir: Path, recipe: str,
                      tmpdir: Path | None = None) -> list[Path]:
    if tmpdir is None:
        tmpdir = _resolve_tmpdir(build_dir)
    lic_dir = tmpdir / "deploy" / "licenses" / recipe
    if not lic_dir.exists():
        return []
    return [f for f in sorted(lic_dir.iterdir())
            if f.is_file()
            and f.name != "recipeinfo"
            and f.suffix.lower() not in _LICENSE_SKIP_EXTS]


def get_patch_files(work_ver: Path) -> list[Path]:
    if not work_ver or not work_ver.exists():
        return []
    return sorted(
        f for f in work_ver.iterdir()
        if f.is_file() and (f.suffix == ".patch" or f.suffix == ".diff")
    )


def parse_src_uri(src_uri: str) -> list[tuple[str, str]]:
    """Parse SRC_URI string into list of (filename, subpath) for ``file://`` entries.

    Skips patches (.patch, .diff) since those are already collected separately.
    Strips SRC_URI parameters (;param=value) before extracting the filename.
    Returns tuples of (bare_filename, original_path_from_uri).
    """
    result: list[tuple[str, str]] = []
    for token in src_uri.split():
        if not token.startswith("file://"):
            continue
        # Strip file:// prefix
        path_part = token[len("file://"):]
        # Strip SRC_URI parameters (;subdir=foo;apply=yes etc.)
        path_part = path_part.split(";")[0]
        if not path_part:
            continue
        # Skip patches
        lower = path_part.lower()
        if lower.endswith(".patch") or lower.endswith(".diff"):
            continue
        filename = os.path.basename(path_part)
        if filename:
            result.append((filename, path_part))
    return result


def get_src_uri_files(work_ver: Path,
                      parsed: list[tuple[str, str]]) -> list[tuple[Path, str]]:
    """Locate actual files from parsed SRC_URI entries in the work directory.

    Returns list of (absolute_path, storage_relative_path) tuples for files
    that exist on disk.
    """
    result: list[tuple[Path, str]] = []
    seen: set[str] = set()
    for filename, subpath in parsed:
        if subpath in seen:
            continue
        seen.add(subpath)
        # Try the subpath directly under work_ver
        candidate = work_ver / subpath
        if candidate.is_file():
            result.append((candidate, subpath))
            continue
        # If subpath has directory components, also try just the filename
        if "/" in subpath:
            candidate = work_ver / filename
            if candidate.is_file():
                result.append((candidate, subpath))
                continue
    return result


def get_pkg_metadata(pkgdata_runtime: Path, yocto_pkg: str) -> dict:
    pkgdata_file = pkgdata_runtime / yocto_pkg
    if not pkgdata_file.exists():
        return {}
    data = parse_pkgdata_file(pkgdata_file)
    result: dict = {}
    for key in ("LICENSE", "HOMEPAGE", "DESCRIPTION", "SUMMARY", "SECTION"):
        if key in data:
            result[key.lower()] = data[key]
    rdep_key = f"RDEPENDS_{yocto_pkg}"
    if rdep_key in data:
        result["rdepends"] = data[rdep_key]
    elif "RDEPENDS" in data:
        result["rdepends"] = data["RDEPENDS"]
    return result


# ── Copyright notice extraction ──

_COPYRIGHT_RE = re.compile(
    r'(?i)copyright\s*(?:\(c\)|©|:)?\s*'
    r'(\d{4}(?:\s*[-–,]\s*\d{4})*)\s+'
    r'(.+?)(?:\.\s*All rights reserved\.?)?$'
)

_COPYRIGHT_SCAN_EXTS = frozenset({
    ".c", ".h", ".cpp", ".cc", ".cxx", ".C", ".S", ".s",
    ".py", ".sh", ".txt", ".md", ".rst", ".pl", ".rb",
})


def extract_copyrights_from_file(path: Path, max_lines: int = 80) -> list[dict]:
    """Scan first max_lines lines of a file for copyright notices."""
    results: list[dict] = []
    try:
        with open(path, "r", errors="replace") as fh:
            for i, line in enumerate(fh):
                if i >= max_lines:
                    break
                m = _COPYRIGHT_RE.search(line)
                if m:
                    holder = m.group(2).strip().rstrip("*/ \t")
                    if holder and len(holder) > 1:
                        results.append({"year": m.group(1).strip(), "holder": holder})
    except (OSError, UnicodeDecodeError):
        pass
    return results


def extract_copyrights_for_recipe(
    sources_dir: Path, licenses_dir: Path, recipe: str,
    pkg_names: list[str],
) -> list[dict]:
    """Extract deduplicated copyright notices for a recipe from collected sources and licenses."""
    seen: set[tuple[str, str]] = set()
    results: list[dict] = []

    scan_dirs: list[Path] = []
    for name in pkg_names:
        d = sources_dir / name
        if d.exists():
            scan_dirs.append(d)
    lic_dir = licenses_dir / recipe
    if lic_dir.exists():
        scan_dirs.append(lic_dir)

    for scan_dir in scan_dirs:
        for fpath in scan_dir.rglob("*"):
            if not fpath.is_file():
                continue
            if fpath.suffix.lower() not in _COPYRIGHT_SCAN_EXTS:
                continue
            for cr in extract_copyrights_from_file(fpath):
                key = (cr["year"], cr["holder"])
                if key not in seen:
                    seen.add(key)
                    results.append(cr)

    results.sort(key=lambda c: (c["holder"].lower(), c["year"]))
    return results


def build_shlibs_map(build_dir: Path, machine: str,
                     tmpdir: Path | None = None) -> dict[str, str]:
    if tmpdir is None:
        tmpdir = _resolve_tmpdir(build_dir)
    shlibs_dir = tmpdir / "pkgdata" / machine / "shlibs2"
    mapping: dict[str, str] = {}
    if not shlibs_dir.exists():
        return mapping
    for list_file in shlibs_dir.iterdir():
        if not list_file.is_file():
            continue
        provider_pkg = list_file.stem
        try:
            for line in list_file.read_text(errors="replace").splitlines():
                line = line.strip()
                if not line:
                    continue
                parts = line.split(":")
                soname = parts[0].strip()
                if soname:
                    mapping[soname] = provider_pkg
        except OSError:
            pass
    return mapping


def get_needed_libs(elf_path: Path) -> list[str]:
    try:
        r = subprocess.run(
            ["readelf", "-d", str(elf_path)],
            capture_output=True, text=True, timeout=30,
        )
    except (subprocess.TimeoutExpired, FileNotFoundError):
        return []
    needed: list[str] = []
    for line in r.stdout.splitlines():
        if "(NEEDED)" in line:
            m = re.search(r'\[(.+?)\]', line)
            if m:
                needed.append(m.group(1))
    return needed


def resolve_linking_deps(
    installed_files: list[dict],
    pkg_split: Path | None,
    shlibs_map: dict[str, str],
    pkgdata_runtime: Path,
    license_cache: dict[str, str],
) -> dict:
    needed_libs: list[dict] = []
    seen_libs: set[str] = set()
    linking_chain: dict[str, str] = {}

    if not pkg_split or not pkg_split.exists():
        return {"needed_libs": needed_libs, "linking_chain": linking_chain}

    for f in installed_files:
        if not f["is_elf"]:
            continue
        elf_path = pkg_split / f["path"].lstrip("/")
        if not elf_path.exists() or elf_path.is_symlink():
            continue
        for lib in get_needed_libs(elf_path):
            if lib in seen_libs:
                continue
            seen_libs.add(lib)
            provider_pkg = shlibs_map.get(lib, "")
            provider_recipe = ""
            provider_license = ""
            copyleft = False
            if provider_pkg:
                if provider_pkg not in license_cache:
                    meta = get_pkg_metadata(pkgdata_runtime, provider_pkg)
                    license_cache[provider_pkg] = meta.get("license", "")
                provider_license = license_cache.get(provider_pkg, "")
                pdata = pkgdata_runtime / provider_pkg
                if pdata.exists():
                    d = parse_pkgdata_file(pdata)
                    provider_recipe = d.get("PN", provider_pkg)
                copyleft = is_copyleft(provider_license) if provider_license else False
                if provider_recipe and provider_license:
                    linking_chain[provider_recipe] = provider_license
            needed_libs.append({
                "lib": lib,
                "provider_pkg": provider_pkg,
                "provider_recipe": provider_recipe,
                "provider_license": provider_license,
                "copyleft": copyleft,
            })

    return {"needed_libs": needed_libs, "linking_chain": linking_chain}


# ── Shared argparse helpers ─────────────────────────────────────────────────

def add_common_args(parser) -> None:
    parser.add_argument(
        "-b", "--build-dir",
        metavar="DIR",
        default=os.environ.get("BUILDDIR", ""),
        help="Yocto build directory (default: $BUILDDIR env var)",
    )
    parser.add_argument(
        "-m", "--manifest",
        metavar="IMAGE_OR_PATH",
        required=True,
        help="Rootfs manifest: image name (e.g. core-image-minimal) or path to .manifest file",
    )
    parser.add_argument(
        "--machine",
        metavar="MACHINE",
        default="",
        help="Yocto MACHINE name (default: auto-detect from pkgdata)",
    )
    parser.add_argument(
        "-v", "--verbose",
        action="store_true",
        help="Show discovery warnings and extra detail",
    )



# ═══════════════════════════════════════════════════════════════════════════════
# Compile-log parsing, coverage checking, and compile testing
# ═══════════════════════════════════════════════════════════════════════════════

def print_list(label: str, items: list[str], limit: int = 8) -> None:
    print(f"  {label} ({len(items)}):")
    for item in items[:limit]:
        print(f"    {item}")
    if len(items) > limit:
        print(f"    … and {len(items) - limit} more")


def _get_initial_cwd(work_ver: Path) -> Path:
    run_script = work_ver / "temp" / "run.do_compile"
    if run_script.exists():
        for line in run_script.read_text(errors="replace").splitlines():
            m = re.match(r"^cd '(.+)'", line)
            if m:
                p = Path(m.group(1))
                if p.is_dir():
                    return p
    return work_ver


def parse_compile_log(log_file: Path, initial_cwd: Path) -> list[CompileCmd]:
    cmds: list[CompileCmd] = []
    try:
        lines = log_file.read_text(errors="replace").splitlines()
    except OSError:
        return cmds

    cwd_stack: dict[int, Path] = {0: initial_cwd}
    current_depth = 0

    for line in lines:
        m = _DEPTH_RE.search(line)
        if m:
            depth  = int(m.group(1))
            action = m.group(2)
            path   = m.group(3)
            if action == "Entering":
                cwd_stack[depth] = Path(path)
                current_depth = depth
            else:
                cwd_stack.pop(depth, None)
                remaining = [k for k in cwd_stack if k < depth]
                current_depth = max(remaining) if remaining else 0
            continue

        line = re.sub(r'^\[\d+/\d+\]\s+', '', line)

        if " -c " not in line or not _COMPILER_RE.search(line):
            continue
        if "--mode=compile" in line:
            continue
        parts = re.split(r'\S*libtool:\s+compile:\s+', line)
        if len(parts) > 1:
            line = parts[-1]
        line = re.sub(r'\s*\|\|\s*\(.*\)\s*$', '', line)
        line = re.sub(r'\s*&&\s*(true|:)\s*$', '', line)
        line = re.sub(r'(?:\s+[12]?>\s*/dev/null|\s+2>&1)+\s*$', '', line)

        cwd = cwd_stack.get(current_depth, initial_cwd)
        cmd = _parse_one_cmd(line, cwd)
        if cmd:
            cmds.append(cmd)

    return cmds


def _parse_one_cmd(line: str, cwd: Path) -> CompileCmd | None:
    try:
        tokens = shlex.split(line)
    except ValueError:
        return None

    obj_tok: str | None = None
    src_tok: str | None = None

    i = 0
    while i < len(tokens):
        if tokens[i] == "-o" and i + 1 < len(tokens):
            obj_tok = tokens[i + 1]
            i += 2
        else:
            i += 1

    for t in reversed(tokens):
        if t.startswith("-"):
            continue
        if "`" in t:
            continue
        if any(t.endswith(e) for e in _COMPILED_EXTS):
            src_tok = t
            break

    if not src_tok:
        return None

    def _abs(tok: str) -> Path:
        p = Path(tok)
        return p if p.is_absolute() else (cwd / p)

    src = _abs(src_tok).resolve()
    obj = _abs(obj_tok).resolve() if obj_tok else None
    return CompileCmd(cwd=cwd, cmd=line.strip(), src=src, obj=obj)


def check_coverage(pkg, sources_dir: Path,
                    all_pkg_names: list[str] | None = None) -> dict:
    """Check how many compiled sources are in the collected output.

    When *all_pkg_names* is given, the check looks for each source in
    ALL listed package directories (recipe-level coverage).  Otherwise
    only the single ``pkg.installed_name`` directory is checked.
    """
    if not pkg.work_ver:
        return {"status": "NO_WORK_DIR"}
    log_file = pkg.work_ver / "temp" / "log.do_compile"
    if not log_file.exists():
        return {"status": "NO_LOG"}
    initial_cwd = _get_initial_cwd(pkg.work_ver)
    cmds = parse_compile_log(log_file, initial_cwd)
    if not cmds:
        return {"status": "NO_CMDS"}

    # Build the list of directories to search for collected sources.
    if all_pkg_names:
        src_dirs = [sources_dir / n for n in all_pkg_names]
    else:
        src_dirs = [sources_dir / pkg.installed_name]

    covered:   list[str] = []
    not_coll:  list[str] = []
    not_coll_abs: dict[str, Path] = {}  # stripped -> absolute source path
    outside:   list[str] = []

    for cmd in cmds:
        try:
            rel = cmd.src.relative_to(pkg.work_ver)
        except ValueError:
            if cmd.src.exists():
                outside.append(str(cmd.src))
            continue
        stripped = strip_src_root(str(rel))
        if any((d / stripped).exists() for d in src_dirs):
            covered.append(stripped)
        else:
            not_coll.append(stripped)
            if cmd.src.exists():
                not_coll_abs[stripped] = cmd.src

    total = len(cmds)
    return {
        "status":   "OK" if not not_coll else "INCOMPLETE",
        "total":    total,
        "covered":  len(covered),
        "not_collected": not_coll,
        "not_collected_abs": not_coll_abs,
        "outside":  outside,
    }


_WORKDIR_INFRA = frozenset((
    "temp", "packages-split", "package", "pkgdata", "license-destdir",
    "sysroot-destdir", "pseudo", "image", "recipe-sysroot",
    "recipe-sysroot-native", "pkgdata-pdata-input", "pkgdata-sysroot",
    "source-date-epoch", "configure.sstate",
    "deploy-rpms", "deploy-ipks", "deploy-debs",
))


def _find_src_subdir(work_ver: Path, recipe: str,
                     ver: str = "",
                     bitbake_s: Path | None = None) -> Path | None:
    """Find the source subdirectory (S) inside a recipe work directory.

    If *bitbake_s* is provided (from ``bitbake -e <recipe>``), use it
    directly.  Otherwise scans work_ver children, skipping known Yocto
    infrastructure dirs, deploy dirs, sstate dirs, build dirs, and
    non-directory entries.
    Uses *ver* (e.g. ``1_5.30.1-r0``) to construct the expected S name
    ``<BPN>-<PV>`` for precise matching.  Falls back to heuristics when
    the exact name is not found.
    """
    if bitbake_s is not None and bitbake_s.is_dir():
        return bitbake_s
    candidates: list[Path] = []
    try:
        children = list(work_ver.iterdir())
    except OSError:
        return None
    for child in children:
        if not child.is_dir():
            continue
        name = child.name
        if name in _WORKDIR_INFRA:
            continue
        if name.startswith("deploy-") or name.startswith("sstate-"):
            continue
        if is_build_dir(name):
            continue
        candidates.append(child)
    if not candidates:
        # Fallback: parse cd from run.do_compile
        cwd = _get_initial_cwd(work_ver)
        if cwd != work_ver and cwd.is_dir():
            return cwd
        return None
    if len(candidates) == 1:
        return candidates[0]
    # Extract PV from ver (strip epoch prefix "N_" and revision suffix "-rN")
    pv = re.sub(r'^(\d+_)', '', ver)            # strip epoch
    pv = re.sub(r'-r\d+$', '', pv)              # strip revision
    pv = re.sub(r'\+git.*$', '', pv)            # strip +gitAUTOINC...
    # BPN: strip numeric-only suffixes from recipe name (glib-2.0 → glib)
    bpn = re.sub(r'-\d[\d.]*$', '', recipe)
    # Try exact match: <BPN>-<PV>
    if pv:
        exact = work_ver / f"{bpn}-{pv}"
        if exact.is_dir() and exact in candidates:
            return exact
        # Also try with full recipe name
        exact2 = work_ver / f"{recipe}-{pv}"
        if exact2.is_dir() and exact2 in candidates:
            return exact2
    # Prefix match — prefer BPN-prefixed candidate
    for c in candidates:
        if c.name.startswith(bpn + "-") or c.name == bpn:
            return c
    for c in candidates:
        if c.name.startswith(recipe + "-") or c.name == recipe:
            return c
    # Last resort: largest directory (most files = likely source tree)
    return max(candidates, key=lambda c: sum(1 for _ in c.iterdir()))


def _check_bitbake_env(build_dir: Path) -> str | None:
    """Validate that bitbake is available. Returns error string or None.

    Uses the same checks as ``_is_bitbake_available()`` but returns a
    descriptive error message instead of a boolean.
    """
    if not shutil.which("bitbake"):
        return ("bitbake not found on PATH. "
                "Source your Yocto build environment first.")
    if not os.environ.get("BUILDDIR") and not os.environ.get("BB_ENV_PASSTHROUGH_ADDITIONS"):
        return ("Yocto build environment not detected "
                "(neither BUILDDIR nor BB_ENV_PASSTHROUGH_ADDITIONS set). "
                "Source your Yocto build environment first.")
    if not (build_dir / "conf" / "local.conf").exists():
        return f"conf/local.conf not found in {build_dir}"
    return None


def _warn_sstate(packages: list[PackageInfo]) -> bool:
    """Warn if build appears to be sstate-only. Returns True if sstate detected."""
    sourced = [p for p in packages if p.pkg_type != "no_source"]
    missing = [p for p in sourced if p.work_ver is None]
    if missing and len(missing) > len(sourced) * 0.5:
        print(f"\n!! WARNING: {len(missing)}/{len(sourced)} source packages have no work directory.")
        print("   This build appears to use sstate (shared state cache).")
        print("   Source collection requires a from-scratch build.")
        print("   Rebuild with: SSTATE_DIR=\"\" SSTATE_MIRRORS=\"\" bitbake --no-setscene <image>\n")
        return True
    return False


def bitbake_compile_test(
    recipe: str,
    src_base: Path,
    sources_dir: Path,
    pkg_names: list[str],
    build_dir: Path,
    timeout: int = 1800,
    verbose: bool = False,
    yocto_tmpdir: Path | None = None,
    bb_env: dict[str, str] | None = None,
) -> dict:
    """Run 'bitbake <recipe> -c compile -f' with collected sources overlaid.

    Algorithm:
      1. Gather collected source files from output/sources/<pkg>/ dirs.
      2. For each collected file, back up the original in the source tree
         and replace it with the collected copy.
      3. Run 'bitbake <recipe> -c compile -f'.
      4. Restore ALL originals from backup (always, even on failure).

    If the compile succeeds, the collected sources are proven to be valid
    compilable replacements.  Coverage completeness is checked separately
    by Phase 1 (check_coverage / file-count checks).
    """
    # ── Step 1: map collected files by relative path ──
    collected_by_rel: dict[str, Path] = {}  # rel_str -> collected_src
    for pkg_name in pkg_names:
        pkg_src_dir = sources_dir / pkg_name
        if not pkg_src_dir.is_dir():
            continue
        for src_file in pkg_src_dir.rglob("*"):
            if not src_file.is_file():
                continue
            rel = str(src_file.relative_to(pkg_src_dir))
            collected_by_rel[rel] = src_file

    if not collected_by_rel:
        return {
            "status": "SKIP",
            "reason": "no collected source files found",
            "recipe": recipe,
        }

    # ── Step 2: overlay collected files onto source tree ──
    _yocto_tmp = yocto_tmpdir if yocto_tmpdir is not None else _resolve_tmpdir(build_dir)
    replaced: dict[Path, Path] = {}  # tree_file -> backup_path
    newly_created: list[Path] = []   # collected files with no original
    tmpdir = None
    try:
        tmpdir = tempfile.mkdtemp(prefix=f"bbtest_{recipe}_")
        tmp_path = Path(tmpdir)

        for rel, collected_src in collected_by_rel.items():
            dest = src_base / rel
            if dest.exists():
                # Back up original and replace
                backup = tmp_path / rel
                backup.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(dest, backup)
                # Make writable if read-only (e.g. perl sources)
                if not os.access(dest, os.W_OK):
                    dest.chmod(dest.stat().st_mode | 0o200)
                shutil.copy2(collected_src, dest)
                replaced[dest] = backup
            else:
                # Collected file not in original tree — copy it in
                dest.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(collected_src, dest)
                newly_created.append(dest)

        if verbose:
            print(f"    Replaced {len(replaced)} source files with collected copies")
            if newly_created:
                print(f"    Created {len(newly_created)} new files (not in original tree)")

        # ── Step 3: run bitbake compile ──
        env = os.environ.copy()
        # Prefer STAGING_BINDIR_NATIVE from bitbake -e, fallback to hosttools
        hosttools: Path | None = None
        if bb_env and bb_env.get("STAGING_BINDIR_NATIVE"):
            p = Path(bb_env["STAGING_BINDIR_NATIVE"])
            if p.is_dir():
                hosttools = p
        if hosttools is None:
            hosttools = _yocto_tmp / "hosttools"
        if hosttools.is_dir():
            env["PATH"] = str(hosttools) + os.pathsep + env.get("PATH", "")

        cmd = ["bitbake", recipe, "-c", "compile", "-f"]
        if verbose:
            print(f"    Running: {' '.join(cmd)}")

        r = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=timeout,
            cwd=str(build_dir),
            env=env,
        )

        # Bitbake may return non-zero due to config warnings (e.g.
        # BB_DISKMON_DIRS) even when do_compile succeeds.  Check stdout
        # for the definitive task-succeeded message.
        compile_ok = (r.returncode == 0
                      or "do_compile: Succeeded" in r.stdout)
        compile_failed = ("do_compile: Failed" in r.stdout
                          or "ERROR: Task" in r.stdout)
        if compile_failed:
            status = "FAIL"
        elif compile_ok:
            status = "PASS"
        else:
            status = "FAIL"
        return {
            "status": status,
            "reason": "" if status == "PASS" else "bitbake compile failed",
            "recipe": recipe,
            "returncode": r.returncode,
            "stdout": r.stdout,
            "stderr": r.stderr,
            "files_replaced": len(replaced),
        }

    except subprocess.TimeoutExpired:
        return {
            "status": "FAIL",
            "reason": f"bitbake compile timed out after {timeout}s",
            "recipe": recipe,
        }
    except Exception as e:
        return {
            "status": "FAIL",
            "reason": str(e),
            "recipe": recipe,
        }
    finally:
        # ── Step 4: always restore originals ──
        for tree_file, backup in replaced.items():
            try:
                shutil.copy2(backup, tree_file)
            except OSError:
                pass
        for dest in newly_created:
            try:
                dest.unlink(missing_ok=True)
            except OSError:
                pass
        if tmpdir:
            shutil.rmtree(tmpdir, ignore_errors=True)
        # Remove the taint file that 'bitbake -c compile -f' creates.
        # Without this, subsequent recipes that depend on this one will
        # see the taint and try to recompile it again.
        stamps_dir: Path | None = None
        if bb_env and bb_env.get("STAMPS_DIR"):
            p = Path(bb_env["STAMPS_DIR"])
            if p.is_dir():
                stamps_dir = p
        if stamps_dir is None:
            stamps_dir = _yocto_tmp / "stamps"
        if stamps_dir.is_dir():
            for taint in stamps_dir.glob(f"*/{recipe}/*.do_compile.taint"):
                try:
                    taint.unlink()
                except OSError:
                    pass


# ═══════════════════════════════════════════════════════════════════════════════
# YoctoSession — shared state for all subcommands
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class YoctoSession:
    build_dir: Path
    manifest_path: Path
    machine: str
    output_dir: Path
    tmpdir: Path
    verbose: bool = False
    _bb_env: dict[str, str] | None = field(default=None, repr=False)
    _bb_recipe_cache: dict[str, dict[str, str]] = field(default_factory=dict, repr=False)
    _packages: list[PackageInfo] | None = field(default=None, repr=False)

    @classmethod
    def from_args(cls, args) -> "YoctoSession":
        if not args.build_dir:
            raise SystemExit(
                "Build directory not specified. Use -b/--build-dir or set $BUILDDIR."
            )
        build_dir = Path(args.build_dir).resolve()
        if not build_dir.is_dir():
            raise SystemExit(f"Build directory not found: {build_dir}")

        # Try bitbake -e for authoritative path discovery (disk-cached)
        cached_global = _bb_cache_load("_global")
        bb_env = _query_bitbake_global(build_dir)
        if bb_env:
            if cached_global is not None:
                print("(using cached bitbake -e for path discovery)")
            else:
                print("(using bitbake -e for path discovery)")
        elif not _is_bitbake_available(build_dir):
            raise SystemExit(
                "bitbake not found. Source your Yocto build environment first."
            )

        tmpdir = _resolve_tmpdir(build_dir, bb_env)
        machine = args.machine or (
            bb_env.get("MACHINE") if bb_env and bb_env.get("MACHINE")
            else auto_detect_machine(build_dir, tmpdir)
        )
        manifest_path = find_manifest(build_dir, args.manifest, tmpdir)
        output_dir = Path("./output").resolve()
        return cls(
            build_dir=build_dir,
            manifest_path=manifest_path,
            machine=machine,
            output_dir=output_dir,
            tmpdir=tmpdir,
            verbose=getattr(args, "verbose", False),
            _bb_env=bb_env,
        )

    def query_recipe_env(self, recipe: str) -> dict[str, str]:
        """Get per-recipe bitbake vars, cached."""
        if recipe not in self._bb_recipe_cache:
            result = _query_bitbake_recipe(recipe, self.build_dir)
            self._bb_recipe_cache[recipe] = result or {}
        return self._bb_recipe_cache[recipe]

    def discover(self) -> list[PackageInfo]:
        if self._packages is None:
            self._packages = discover_packages(
                self.manifest_path, self.build_dir, self.machine,
                self.verbose, self.tmpdir, session=self,
            )
        return self._packages

    @property
    def sources_dir(self) -> Path:
        return self.output_dir / "sources"

    @property
    def pkgdata_dir(self) -> Path:
        return self.tmpdir / "pkgdata" / self.machine

    @property
    def work_dir(self) -> Path:
        return self.tmpdir / "work"

    def print_header(self) -> None:
        print(f"Build dir : {self.build_dir}")
        print(f"TMPDIR    : {self.tmpdir}")
        print(f"Machine   : {self.machine}")
        print(f"Manifest  : {self.manifest_path}")
        print(f"Output    : {self.output_dir}")


# ═══════════════════════════════════════════════════════════════════════════════
# Collector — collect sources per installed package in three categories
# ═══════════════════════════════════════════════════════════════════════════════



def _resolve_dwarf_source(work_ver: Path, rel: str) -> Path | None:
    """Try multiple strategies to locate a DWARF-referenced source file."""
    # Strategy 1: direct path
    candidate = work_ver / rel
    if candidate.exists():
        return candidate
    # Strategy 2: scan first-level subdirectories (source root may be missing)
    for child in work_ver.iterdir():
        if child.is_dir() and child.name not in _WORKDIR_INFRA:
            candidate = child / rel
            if candidate.exists():
                return candidate
    # Strategy 3: rglob by basename (generated files in build/ dirs, cross-prefix paths)
    basename = Path(rel).name
    if basename:
        for match in work_ver.rglob(basename):
            if match.is_file():
                return match
    return None


def copy_source(src: Path, dst: Path) -> bool:
    if not src.exists():
        return False
    dst.parent.mkdir(parents=True, exist_ok=True)
    if dst.exists():
        dst.chmod(0o644)
    shutil.copy2(src, dst)
    return True


def _get_compiled_srcs(work_ver: Path) -> set[Path]:
    log_file = work_ver / "temp" / "log.do_compile"
    if not log_file.exists():
        return set()
    try:
        cwd = _get_initial_cwd(work_ver)
        cmds = parse_compile_log(log_file, cwd)
        return {cmd.src for cmd in cmds}
    except Exception:
        return set()


def _count_files(d: Path) -> int:
    if not d.exists():
        return 0
    return sum(1 for f in d.rglob("*") if f.is_file())


class Collector:
    def __init__(self, session: YoctoSession, clean: bool = False):
        self.session = session
        self.clean = clean
        self.out_dir = session.sources_dir
        self.licenses_dir = session.output_dir / "licenses"
        self.patches_dir = session.output_dir / "patches"
        self.srcuri_dir = session.output_dir / "src_uri"
        self.copyrights_dir = session.output_dir / "copyrights"

    def run(self) -> list[tuple[str, str]]:
        """Run collection. Returns list of (name, reason) tuples for failures."""
        session = self.session
        session.print_header()

        if self.clean:
            for d in (self.out_dir, self.licenses_dir, self.patches_dir,
                      self.srcuri_dir, self.copyrights_dir, _bb_cache_dir()):
                if d.exists():
                    shutil.rmtree(d)
            print("(cleaned output dirs)")

        self.out_dir.mkdir(parents=True, exist_ok=True)

        packages = session.discover()
        print(f"\nDiscovered {len(packages)} packages\n")
        _warn_sstate(packages)

        kernel_image_done: dict[tuple[str, str], bool] = {}
        recipes_done: set[str] = set()
        total_lic = 0
        total_pat = 0
        total_srcuri = 0
        failed_packages: list[tuple[str, str]] = []

        for pkg in packages:
            print(f"[{pkg.installed_name}]  type={pkg.pkg_type}  recipe={pkg.recipe}  ver={pkg.ver}")

            if pkg.pkg_type == "no_source":
                self.write_no_source(pkg)
                print("  → (no compiled source)")
            elif pkg.pkg_type == "kernel_image":
                key = (pkg.recipe, pkg.ver)
                if not kernel_image_done.get(key):
                    counts = self.collect_kernel_image(pkg)
                    kernel_image_done[key] = True
                    if "error" in counts:
                        print(f"  !! ERROR: {counts['error']}")
                        failed_packages.append((pkg.installed_name, counts["error"]))
                    elif counts["c"] + counts["S"] == 0:
                        reason = self._diagnose_kernel_image_zero(pkg)
                        print(f"  !! FAIL: kernel_image collected 0 source files — {reason}")
                        failed_packages.append((pkg.installed_name, reason))
                    print(f"  → c={counts['c']}  h={counts['h']}  S={counts['S']}  missing={counts['missing']}")
                else:
                    note_dir = self.out_dir / pkg.installed_name
                    note_dir.mkdir(parents=True, exist_ok=True)
                    primary = next(
                        (p.installed_name for p in packages
                         if p.pkg_type == "kernel_image"
                         and p.recipe == pkg.recipe and p.ver == pkg.ver
                         and kernel_image_done.get((p.recipe, p.ver))),
                        "kernel-image package"
                    )
                    (note_dir / f"SAME_AS_{primary}.txt").write_text(
                        f"Sources for {pkg.installed_name} are identical to {primary}.\n"
                        f"See that directory for the full source list.\n"
                    )
                    print(f"  → shares sources with {primary}")
            elif pkg.pkg_type == "kernel_module":
                counts = self.collect_kernel_module(pkg)
                if "error" in counts:
                    print(f"  !! ERROR: {counts['error']}")
                    failed_packages.append((pkg.installed_name, counts["error"]))
                elif counts["c"] + counts["S"] == 0:
                    reason = self._diagnose_kernel_module_zero(pkg)
                    print(f"  !! FAIL: kernel_module collected 0 source files — {reason}")
                    failed_packages.append((pkg.installed_name, reason))
                print(f"  → c={counts['c']}  h={counts.get('h', 0)}  S={counts['S']}  missing={counts['missing']}")
            else:
                counts = self.collect_userspace(pkg)
                print(f"  → sources={counts['compiled_used']}  missing={counts['missing']}")

            # Collect licenses, patches, and SRC_URI files per recipe (deduplicated)
            if pkg.recipe not in recipes_done:
                recipes_done.add(pkg.recipe)
                n_lic = self.collect_licenses(pkg.recipe, session.build_dir, session.tmpdir)
                n_pat = self.collect_patches(pkg.recipe, pkg.work_ver)
                n_uri = self.collect_src_uri_files(pkg.recipe, pkg.work_ver)
                total_lic += n_lic
                total_pat += n_pat
                total_srcuri += n_uri
                if n_lic or n_pat or n_uri:
                    parts = []
                    if n_lic:
                        parts.append(f"licenses={n_lic}")
                    if n_pat:
                        parts.append(f"patches={n_pat}")
                    if n_uri:
                        parts.append(f"yocto_additional_sources={n_uri}")
                    print(f"  → {'  '.join(parts)}")

        self.write_manifest(packages)

        # Collect copyright notices per recipe
        n_copyrights = self.collect_copyrights(packages, recipes_done)

        pkg_dirs = len([d for d in self.out_dir.iterdir() if d.is_dir()])
        print(f"Done. {pkg_dirs} package directories in {self.out_dir}")
        print(f"  Licenses:    {total_lic} files in {self.licenses_dir}")
        print(f"  Patches:     {total_pat} files in {self.patches_dir}")
        print(f"  Yocto additional sources: {total_srcuri} files in {self.srcuri_dir}")
        print(f"  Copyrights:  {n_copyrights} recipe files in {self.copyrights_dir}")

        if failed_packages:
            print(f"\n!! COLLECTION FAILED for {len(failed_packages)} package(s):")
            for name, reason in failed_packages:
                print(f"   - {name}: {reason}")
            print("These packages are kernel/module type and MUST have source files.")

        return failed_packages

    def collect_copyrights(self, packages: list[PackageInfo],
                            recipes_done: set[str]) -> int:
        """Extract copyright notices per recipe and write to output/copyrights/."""
        self.copyrights_dir.mkdir(parents=True, exist_ok=True)
        recipe_pkgs: dict[str, list[str]] = {}
        for pkg in packages:
            if pkg.recipe in recipes_done:
                recipe_pkgs.setdefault(pkg.recipe, []).append(pkg.installed_name)

        count = 0
        for recipe, pkg_names in sorted(recipe_pkgs.items()):
            notices = extract_copyrights_for_recipe(
                self.out_dir, self.licenses_dir, recipe, pkg_names)
            if notices:
                out_file = self.copyrights_dir / f"{recipe}.txt"
                lines = [f"{cr['year']}  {cr['holder']}" for cr in notices]
                out_file.write_text("\n".join(lines) + "\n")
                count += 1
        return count

    def _diagnose_kernel_image_zero(self, pkg: PackageInfo) -> str:
        """Diagnose why collect_kernel_image found 0 .c/.S sources."""
        k = pkg.kernel
        if k is None:
            return "KernelInfo is None (should have been caught earlier)"
        # Check .o files in build dir
        o_files = list(k.build_dir.rglob("*.o"))
        n_o = len(o_files)
        if n_o == 0:
            # List what IS in build dir
            top_entries = sorted(e.name for e in k.build_dir.iterdir())[:10] if k.build_dir.is_dir() else []
            return (f"kernel build dir has 0 .o files ({k.build_dir}). "
                    f"Build artifacts may have been cleaned. "
                    f"Top entries: {top_entries}")
        # .o files exist but no source matched — check src_dir
        if not k.src_dir.is_dir():
            return (f"kernel source dir does not exist: {k.src_dir}")
        # Sample a few .o files and show what source paths were tried
        samples = []
        for o_file in o_files[:3]:
            rel = o_file.relative_to(k.build_dir)
            c_path = k.src_dir / rel.parent / (rel.stem + ".c")
            s_path = k.src_dir / rel.parent / (rel.stem + ".S")
            samples.append(f"{rel} → tried {c_path} ({'exists' if c_path.exists() else 'MISSING'}), "
                           f"{s_path} ({'exists' if s_path.exists() else 'MISSING'})")
        src_entries = sorted(e.name for e in k.src_dir.iterdir())[:10] if k.src_dir.is_dir() else []
        return (f"{n_o} .o files in {k.build_dir} but no matching .c/.S in {k.src_dir}. "
                f"src_dir top entries: {src_entries}. "
                f"Sample lookups: {'; '.join(samples)}")

    def _diagnose_kernel_module_zero(self, pkg: PackageInfo) -> str:
        """Diagnose why collect_kernel_module found 0 .c/.S sources."""
        k = pkg.kernel
        if k is None:
            return "KernelInfo is None (should have been caught earlier)"
        if not pkg.kernel_mod_obj_rels:
            # Check .ko files in packages-split
            pkg_split = pkg.work_ver / "packages-split" / pkg.yocto_pkg if pkg.work_ver else None
            ko_files = list(pkg_split.rglob("*.ko")) if pkg_split and pkg_split.is_dir() else []
            if not ko_files:
                return (f"no .ko files found in packages-split/{pkg.yocto_pkg} "
                        f"— package may not contain kernel modules")
            # .ko exist but no .mod files
            ko_names = [f.name for f in ko_files]
            mod_checked = []
            for ko_name in ko_names[:3]:
                found_in_build = list(k.build_dir.rglob(ko_name))
                for ko_in_build in found_in_build:
                    mod_file = ko_in_build.parent / (ko_in_build.stem + ".mod")
                    mod_checked.append(f"{mod_file} ({'exists' if mod_file.exists() else 'MISSING'})")
            return (f".ko files present ({ko_names}) but no .mod files found to resolve .o objects. "
                    f"Checked: {'; '.join(mod_checked) if mod_checked else 'no matching .ko in build dir'}")
        # obj_rels exist but no source matched
        samples = []
        for obj_rel in pkg.kernel_mod_obj_rels[:3]:
            rel = Path(obj_rel)
            c_path = k.src_dir / rel.parent / (rel.stem + ".c")
            s_path = k.src_dir / rel.parent / (rel.stem + ".S")
            samples.append(f"{rel} → tried {c_path} ({'exists' if c_path.exists() else 'MISSING'}), "
                           f"{s_path} ({'exists' if s_path.exists() else 'MISSING'})")
        return (f"{len(pkg.kernel_mod_obj_rels)} .o rel(s) resolved but no matching .c/.S in {k.src_dir}. "
                f"Sample lookups: {'; '.join(samples)}")

    def collect_licenses(self, recipe: str, build_dir: Path,
                         tmpdir: Path | None = None) -> int:
        files = get_license_files(build_dir, recipe, tmpdir)
        if not files:
            return 0
        dest = self.licenses_dir / recipe
        dest.mkdir(parents=True, exist_ok=True)
        count = 0
        for f in files:
            dst = dest / f.name
            if not dst.exists():
                shutil.copy2(f, dst)
                count += 1
        return count

    def collect_patches(self, recipe: str, work_ver: Path | None) -> int:
        files = get_patch_files(work_ver)
        if not files:
            return 0
        dest = self.patches_dir / recipe
        dest.mkdir(parents=True, exist_ok=True)
        count = 0
        for f in files:
            dst = dest / f.name
            if not dst.exists():
                shutil.copy2(f, dst)
                count += 1
        return count

    def collect_src_uri_files(self, recipe: str,
                              work_ver: Path | None) -> int:
        """Collect non-patch file:// SRC_URI entries for a recipe."""
        if not work_ver or not work_ver.exists():
            return 0
        renv = self.session.query_recipe_env(recipe)
        src_uri = renv.get("SRC_URI", "")
        if not src_uri:
            return 0
        parsed = parse_src_uri(src_uri)
        if not parsed:
            return 0
        files = get_src_uri_files(work_ver, parsed)
        if not files:
            return 0
        dest = self.srcuri_dir / recipe
        dest.mkdir(parents=True, exist_ok=True)
        count = 0
        for abs_path, rel_path in files:
            dst = dest / rel_path
            if not dst.exists():
                dst.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(abs_path, dst)
                count += 1
        return count

    def collect_userspace(self, pkg_info: PackageInfo) -> dict:
        build_dir = self.session.build_dir
        tmpdir = self.session.tmpdir
        split_root = pkg_info.work_ver / "packages-split"
        pkg_split  = split_root / pkg_info.yocto_pkg
        prefix     = pkg_info.recipe_prefix

        counts = {"compiled_used": 0, "missing": 0}
        pkg_out = self.out_dir / pkg_info.installed_name

        compiled_srcs: set[Path] = _get_compiled_srcs(pkg_info.work_ver)

        installed_elfs = find_installed_elfs(pkg_split)
        if installed_elfs:
            dwarf_items: list[tuple[Path, str]] = []
            dwarf_abs:   set[Path] = set()

            for elf in installed_elfs:
                dbg = find_debug_counterpart(elf, pkg_split)
                target = dbg if (dbg and dbg.exists()) else elf
                for path in extract_dwarf_cu_sources(target):
                    if path.startswith(prefix):
                        rel = path[len(prefix):]
                        if rel and not rel.startswith("<"):
                            src = _resolve_dwarf_source(pkg_info.work_ver, rel) or (pkg_info.work_ver / rel)
                            dst = strip_src_root(rel)
                            if src not in dwarf_abs:
                                dwarf_items.append((src, dst))
                                dwarf_abs.add(src)
                    elif build_dir and "/work-shared/" in path:
                        ws_idx = path.find("/work-shared/")
                        ws_rel = path[ws_idx + 1:]
                        src = tmpdir / ws_rel
                        after_ws = ws_rel.split("/", 1)[1] if "/" in ws_rel else ""
                        dst = strip_src_root(after_ws) if after_ws else ""
                        if dst and not dst.startswith("<") and src not in dwarf_abs:
                            dwarf_items.append((src, dst))
                            dwarf_abs.add(src)

            if dwarf_items:
                for src, dst_rel in dwarf_items:
                    if copy_source(src, pkg_out / dst_rel):
                        counts["compiled_used"] += 1
                    else:
                        counts["missing"] += 1

                debugsources = pkg_info.work_ver / "debugsources.list"
                if debugsources.exists():
                    for debug_path in read_debugsources(debugsources):
                        src_h = None
                        dst_rel_h = None
                        if debug_path.startswith(prefix):
                            rel = debug_path[len(prefix):]
                            if not rel or rel.startswith("<") or not rel.endswith(".h"):
                                continue
                            src_h = _resolve_dwarf_source(pkg_info.work_ver, rel) or (pkg_info.work_ver / rel)
                            dst_rel_h = strip_src_root(rel)
                        elif build_dir and "/work-shared/" in debug_path:
                            if not debug_path.endswith(".h"):
                                continue
                            ws_idx = debug_path.find("/work-shared/")
                            ws_rel = debug_path[ws_idx + 1:]
                            src_h = tmpdir / ws_rel
                            after_ws = ws_rel.split("/", 1)[1] if "/" in ws_rel else ""
                            dst_rel_h = strip_src_root(after_ws) if after_ws else ""
                        if src_h and dst_rel_h:
                            if copy_source(src_h, pkg_out / dst_rel_h):
                                counts["compiled_used"] += 1

                return counts

            # No DWARF data but compile log has entries — collect from log
            if compiled_srcs:
                for abs_src in compiled_srcs:
                    try:
                        rel = str(abs_src.relative_to(pkg_info.work_ver))
                    except ValueError:
                        continue
                    dst_rel = strip_src_root(rel)
                    if copy_source(abs_src, pkg_out / dst_rel):
                        counts["compiled_used"] += 1

                return counts

        # Fallback: recipe-level debugsources.list
        debugsources = pkg_info.work_ver / "debugsources.list"
        paths = read_debugsources(debugsources)

        for debug_path in paths:
            if not debug_path.startswith(prefix):
                continue
            rel = debug_path[len(prefix):]
            if not rel or rel.startswith("<"):
                continue
            src     = pkg_info.work_ver / rel
            dst_rel = strip_src_root(rel)
            if copy_source(src, pkg_out / dst_rel):
                counts["compiled_used"] += 1
            else:
                counts["missing"] += 1

        return counts

    @staticmethod
    def _kernel_no_info_error(pkg_info: PackageInfo) -> dict:
        """Return error dict when KernelInfo is unavailable."""
        if not pkg_info.work_ver:
            reason = "work_ver is None — sstate-only build, no work directory found"
        else:
            kbuild = find_kernel_build_dir(pkg_info.work_ver)
            if not kbuild:
                reason = (f"no *-build directory found in {pkg_info.work_ver} "
                          f"— kernel may not have been compiled, or build artifacts were cleaned")
            else:
                reason = "kernel-source not found — work-shared kernel source missing or cleaned"
        return {"c": 0, "S": 0, "h": 0, "missing": 0, "error": reason}

    def collect_kernel_image(self, pkg_info: PackageInfo) -> dict:
        k = pkg_info.kernel
        if k is None:
            return self._kernel_no_info_error(pkg_info)

        module_objs = k.module_objs()
        pkg_out = self.out_dir / pkg_info.installed_name
        counts = {"c": 0, "S": 0, "h": 0, "missing": 0}

        src_prefix = str(k.src_dir)
        collected_h: set[str] = set()
        n_o = 0

        for o_file in k.build_dir.rglob("*.o"):
            if o_file in module_objs:
                continue
            if o_file.name.endswith(".mod.o"):
                continue
            n_o += 1
            rel = o_file.relative_to(k.build_dir)

            # Collect .c/.S source for this object
            for ext in (".c", ".S"):
                src = k.src_dir / rel.parent / (rel.stem + ext)
                if copy_source(src, pkg_out / rel.parent / src.name):
                    counts["c" if ext == ".c" else "S"] += 1
                    break

            # Collect .h headers from Kbuild .cmd dependency file
            cmd_file = o_file.parent / f".{o_file.name}.cmd"
            if not cmd_file.exists():
                continue
            try:
                text = cmd_file.read_text(errors="replace")
            except OSError:
                continue
            for line in text.splitlines():
                h = line.strip().rstrip("\\").strip()
                if not h.endswith(".h") or not h.startswith(src_prefix):
                    continue
                if h in collected_h:
                    continue
                h_path = Path(h)
                try:
                    rel_h = h_path.relative_to(k.src_dir)
                except ValueError:
                    continue
                if copy_source(h_path, pkg_out / rel_h):
                    collected_h.add(h)
                    counts["h"] += 1

        total_sources = counts["c"] + counts["S"]
        if total_sources == 0:
            if n_o == 0:
                counts["error"] = (f"kernel build dir has no .o files: {k.build_dir} "
                                   f"— build artifacts may have been cleaned (do_clean/sstate)")
            else:
                counts["error"] = (f"{n_o} .o files found but no matching sources in "
                                   f"{k.src_dir} — kernel-source may be incomplete or path mismatch")

        return counts

    def collect_kernel_module(self, pkg_info: PackageInfo) -> dict:
        k = pkg_info.kernel
        if k is None:
            return self._kernel_no_info_error(pkg_info)

        pkg_out = self.out_dir / pkg_info.installed_name
        counts = {"c": 0, "S": 0, "h": 0, "missing": 0}

        for obj_rel in pkg_info.kernel_mod_obj_rels:
            rel = Path(obj_rel)
            found = False
            for ext in (".c", ".S"):
                src = k.src_dir / rel.parent / (rel.stem + ext)
                if copy_source(src, pkg_out / rel.parent / src.name):
                    counts["c" if ext == ".c" else "S"] += 1
                    found = True
                    break
            if not found:
                counts["missing"] += 1

        src_prefix = str(k.src_dir)
        collected_h: set[str] = set()

        for obj_rel in pkg_info.kernel_mod_obj_rels:
            o_file = k.build_dir / obj_rel
            cmd_file = o_file.parent / f".{o_file.name}.cmd"
            if not cmd_file.exists():
                continue
            try:
                text = cmd_file.read_text(errors="replace")
            except OSError:
                continue
            for line in text.splitlines():
                h = line.strip().rstrip("\\").strip()
                if not h.endswith(".h") or not h.startswith(src_prefix):
                    continue
                if h in collected_h:
                    continue
                h_path = Path(h)
                try:
                    rel_h = h_path.relative_to(k.src_dir)
                except ValueError:
                    continue
                if copy_source(h_path, pkg_out / rel_h):
                    collected_h.add(h)
                    counts["h"] += 1

        if not pkg_info.kernel_mod_obj_rels:
            counts["error"] = (f"no .mod files found for this module's .ko — "
                               f"cannot determine which .o files belong to this module. "
                               f"Kernel build may have been partially cleaned")

        pkg_out.mkdir(parents=True, exist_ok=True)
        return counts

    def write_no_source(self, pkg_info: PackageInfo) -> None:
        pkg_out = self.out_dir / pkg_info.installed_name
        pkg_out.mkdir(parents=True, exist_ok=True)

    def write_manifest(self, packages: list[PackageInfo]) -> None:
        rows = []
        for p in sorted(packages, key=lambda x: x.installed_name):
            d = self.out_dir / p.installed_name
            cu = _count_files(d)
            rows.append((p.installed_name, p.recipe, p.ver, cu, str(d)))

        manifest = self.out_dir / "MANIFEST.txt"
        header = (f"{'package':<55} {'recipe':<20} {'version':<45}"
                  f" {'source_files':>12}  source_dir\n")
        sep = "-" * 145 + "\n"
        lines = [header, sep]
        for name, rec, ver, cu, sdir in rows:
            lines.append(
                f"{name:<55} {rec:<20} {ver:<45}"
                f" {cu:>12}  {sdir}\n"
            )
        manifest.write_text("".join(lines))
        print(f"\nManifest: {manifest}")


# ═══════════════════════════════════════════════════════════════════════════════
# Verifier — cross-check collected sources against installed rootfs binaries
# ═══════════════════════════════════════════════════════════════════════════════

class Verifier:
    def __init__(self, session: YoctoSession):
        self.session = session
        self.out_dir = session.sources_dir

    def run(self) -> int:
        session = self.session
        session.print_header()
        print(f"Sources   : {self.out_dir}")

        if not self.out_dir.exists():
            raise SystemExit(f"Sources directory not found: {self.out_dir}\n"
                             "Run 'python3 yocto/source_audit.py collect' first.")

        packages = session.discover()
        print(f"\nDiscovered {len(packages)} packages\n")
        print("=" * 72)

        results: dict[str, str] = {}
        kernel_image_verified: dict[tuple[str, str], str] = {}

        for pkg in sorted(packages, key=lambda p: p.installed_name):
            print(f"\n[{pkg.installed_name}]")

            if pkg.pkg_type == "no_source":
                r = self.verify_no_source(pkg)
                print(f"  {'OK' if r['status']=='OK' else 'FAIL'}: package dir "
                      f"{'present' if r['status']=='OK' else 'MISSING'} (empty — no compiled source)")
                results[pkg.installed_name] = r["status"]
                continue

            if pkg.pkg_type == "kernel_image":
                key = (pkg.recipe, pkg.ver)
                if key not in kernel_image_verified:
                    r = self.verify_kernel_image(pkg)
                    kernel_image_verified[key] = pkg.installed_name
                    print("  Method: .o file cross-check (vmlinux has no DWARF)")
                    print(f"  Collected: {r.get('total_collected', 0)}")
                    print(f"  Verified (non-module .o present): {r.get('verified', 0)}")
                    if r.get("missing_o"):
                        print_list("Collected but .o missing", r["missing_o"])
                    if r.get("is_module_src"):
                        print_list("Collected but .o is module-owned", r["is_module_src"])
                    if r.get("uncollected"):
                        print_list("Non-module .o exists but NOT collected", r["uncollected"])
                    if r["status"] == "OK":
                        print("  PERFECT MATCH")
                    results[pkg.installed_name] = r["status"]
                else:
                    primary = kernel_image_verified[key]
                    note = self.out_dir / pkg.installed_name / f"SAME_AS_{primary}.txt"
                    ok = note.exists() or (self.out_dir / pkg.installed_name).exists()
                    print(f"  Shares sources with {primary} — {'OK' if ok else 'FAIL'}")
                    results[pkg.installed_name] = "OK" if ok else "FAIL"
                continue

            if pkg.pkg_type == "kernel_module":
                r = self.verify_kernel_module(pkg)
                print(f"  Expected (from .mod): {r.get('expected', [])}")
                print(f"  Collected:            {r.get('collected', [])}")
                if r.get("missing"): print_list("MISSING", r["missing"])
                if r.get("extra"):   print_list("EXTRA",   r["extra"])
                if r["status"] == "OK": print("  PERFECT MATCH")
                results[pkg.installed_name] = r["status"]
                continue

            r = self.verify_userspace(pkg)
            status = r.get("status", "ERROR")
            results[pkg.installed_name] = status

            if status == "NO_BINARIES":
                print(f"  WARN: no ELFs in packages-split/{pkg.yocto_pkg}/")
            elif status == "NO_DWARF_SOURCES":
                print(f"  WARN: no own-recipe DWARF CUs "
                      f"({r['elfs']} ELF(s), {r['debug_found']} with .debug)")
                print(f"  Collected: {len(list_collected_files(self.out_dir, pkg.installed_name))} files")
            else:
                dbg = r.get("debug_found", 0)
                total = len(r.get("binaries", []))
                print(f"  Binaries: {', '.join(r['binaries'])}  ({dbg}/{total} with .debug)")
                print(f"  DWARF sources (own recipe): {r['dwarf']}")
                print(f"  Collected:                  {r['collected']}")
                if r.get("missing"):
                    print_list("MISSING from sources/", r["missing"])
                if r.get("extra_c"):
                    print_list("EXTRA .c files (not in DWARF)", r["extra_c"])
                if not r.get("missing") and not r.get("extra_c"):
                    print("  PERFECT MATCH  (.h extras are expected — DWARF CU-level only)")

        print("\n" + "=" * 72)
        print("SUMMARY")
        print("=" * 72)

        ok_pkgs   = [p for p, s in results.items() if s == "OK"]
        warn_pkgs = [p for p, s in results.items()
                     if s in ("NO_DWARF_SOURCES", "NO_BINARIES", "MISSING",
                               "ISSUES", "MISMATCH")]
        fail_pkgs = [p for p, s in results.items() if s == "FAIL"]

        print(f"  OK:   {len(ok_pkgs)}")
        print(f"  WARN: {len(warn_pkgs)}")
        for p in warn_pkgs:
            print(f"         {p}  [{results[p]}]")
        print(f"  FAIL: {len(fail_pkgs)}  {fail_pkgs}")

        return 0 if not fail_pkgs else 1

    def verify_no_source(self, pkg_info: PackageInfo) -> dict:
        exists = (self.out_dir / pkg_info.installed_name).is_dir()
        return {"status": "OK" if exists else "FAIL",
                "detail": "" if exists else "package directory missing"}

    def verify_userspace(self, pkg_info: PackageInfo) -> dict:
        split_root = pkg_info.work_ver / "packages-split"
        pkg_split  = split_root / pkg_info.yocto_pkg
        prefix     = pkg_info.recipe_prefix

        installed_elfs = find_installed_elfs(pkg_split)
        if not installed_elfs:
            return {"status": "NO_BINARIES", "pkg_split": str(pkg_split)}

        dwarf_sources: set[str] = set()
        debug_found = 0

        for elf in installed_elfs:
            dbg = find_debug_counterpart(elf, pkg_split)
            target = dbg if (dbg and dbg.exists()) else elf
            if dbg and dbg.exists():
                debug_found += 1
            for path in extract_dwarf_cu_sources(target):
                if path.startswith(prefix):
                    rel = path[len(prefix):]
                    if rel and not rel.startswith("<"):
                        dwarf_sources.add(strip_src_root(rel))

        if not dwarf_sources:
            return {
                "status": "NO_DWARF_SOURCES",
                "elfs": len(installed_elfs),
                "debug_found": debug_found,
            }

        collected = list_collected_files(self.out_dir, pkg_info.installed_name)
        missing   = dwarf_sources - collected
        extra_c   = {f for f in (collected - dwarf_sources) if f.endswith(".c")}

        return {
            "status":      "OK" if not missing else "MISSING",
            "dwarf":       len(dwarf_sources),
            "collected":   len(collected),
            "missing":     sorted(missing),
            "extra_c":     sorted(extra_c),
            "binaries":    [e.name for e in installed_elfs],
            "debug_found": debug_found,
        }

    def verify_kernel_module(self, pkg_info: PackageInfo) -> dict:
        k = pkg_info.kernel
        if k is None:
            return {"status": "NO_KERNEL_INFO"}

        expected: set[str] = set()
        for obj_rel in pkg_info.kernel_mod_obj_rels:
            rel = Path(obj_rel)
            for ext in (".c", ".S"):
                src = k.src_dir / rel.parent / (rel.stem + ext)
                if src.exists():
                    expected.add(str(src.relative_to(k.src_dir)))
                    break

        collected = list_collected_files(self.out_dir, pkg_info.installed_name)
        missing = expected  - collected
        extra   = collected - expected

        return {
            "status":    "OK" if not missing and not extra else "MISMATCH",
            "expected":  sorted(expected),
            "collected": sorted(collected),
            "missing":   sorted(missing),
            "extra":     sorted(extra),
        }

    def verify_kernel_image(self, pkg_info: PackageInfo) -> dict:
        k = pkg_info.kernel
        if k is None:
            return {"status": "NO_KERNEL_INFO"}

        src_dir = self.out_dir / pkg_info.installed_name
        if not src_dir.exists():
            return {"status": "NO_SOURCES_DIR", "detail": str(src_dir)}

        module_objs = k.module_objs()

        missing_o:     list[str] = []
        is_module_src: list[str] = []
        uncollected:   list[str] = []
        verified = 0

        for src_file in src_dir.rglob("*"):
            if not src_file.is_file() or src_file.suffix not in (".c", ".S", ".s"):
                continue
            rel  = src_file.relative_to(src_dir)
            o    = k.build_dir / rel.parent / (rel.stem + ".o")
            if o in module_objs:
                is_module_src.append(str(rel))
            elif o.exists():
                verified += 1
            else:
                missing_o.append(str(rel))

        for o_file in k.build_dir.rglob("*.o"):
            if o_file in module_objs:
                continue
            if o_file.name.startswith(".") or o_file.name.endswith(".mod.o"):
                continue
            rel = o_file.relative_to(k.build_dir)
            collected_any = any(
                (src_dir / rel.parent / (rel.stem + ext)).exists()
                for ext in (".c", ".S", ".s")
            )
            if not collected_any:
                for ext in (".c", ".S", ".s"):
                    ks = k.src_dir / rel.parent / (rel.stem + ext)
                    if ks.exists():
                        uncollected.append(str(rel.parent / (rel.stem + ext)))
                        break

        total = verified + len(missing_o) + len(is_module_src)
        ok = not missing_o and not is_module_src and not uncollected
        return {
            "status":          "OK" if ok else "ISSUES",
            "total_collected": total,
            "verified":        verified,
            "missing_o":       missing_o[:20],
            "is_module_src":   is_module_src[:20],
            "uncollected":     uncollected[:20],
        }


# ═══════════════════════════════════════════════════════════════════════════════
# HTML template for the interactive source report
# ═══════════════════════════════════════════════════════════════════════════════

HTML_TEMPLATE = """\
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Yocto Source Report — {image}</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4/dist/chart.umd.min.js"></script>
<style>
:root{{
  --bg:#f5f7fa;--card:#fff;--border:#e0e4ea;
  --text:#1a1d23;--muted:#6b7280;--accent:#4f86c6;
  --rad:10px;--shadow:0 2px 8px rgba(0,0,0,.08);
  --cu:#27ae60;
}}
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:system-ui,sans-serif;background:var(--bg);color:var(--text);font-size:14px;line-height:1.5}}
header{{background:#1a1d23;color:#fff;padding:20px 32px}}
header h1{{font-size:1.4rem;font-weight:600}}
header p{{color:#9ca3af;font-size:.85rem;margin-top:4px}}
.wrap{{max-width:1600px;margin:0 auto;padding:24px 24px 60px}}

/* Cards */
.cards{{display:grid;grid-template-columns:repeat(auto-fit,minmax(140px,1fr));gap:14px;margin-bottom:24px}}
.card{{background:var(--card);border:1px solid var(--border);border-radius:var(--rad);padding:16px 18px;box-shadow:var(--shadow)}}
.card .val{{font-size:1.8rem;font-weight:700}}
.card .lbl{{color:var(--muted);font-size:.78rem;margin-top:2px}}
.c3 .val{{color:var(--cu)}}.c4 .val{{color:#e07b39}}

/* Charts */
.charts{{display:grid;grid-template-columns:1fr 260px;gap:18px;margin-bottom:24px}}
@media(max-width:800px){{.charts{{grid-template-columns:1fr}}}}
.cc{{background:var(--card);border:1px solid var(--border);border-radius:var(--rad);padding:18px;box-shadow:var(--shadow)}}
.cc h2{{font-size:.85rem;font-weight:600;color:var(--muted);margin-bottom:12px;text-transform:uppercase;letter-spacing:.05em}}

/* Table */
.tbl-wrap{{background:var(--card);border:1px solid var(--border);border-radius:var(--rad);box-shadow:var(--shadow);overflow:hidden}}
.toolbar{{display:flex;align-items:center;gap:10px;padding:12px 16px;border-bottom:1px solid var(--border);flex-wrap:wrap}}
.toolbar h2{{font-size:.85rem;font-weight:600;color:var(--muted);text-transform:uppercase;letter-spacing:.05em;flex:1}}
.toolbar input,.toolbar select{{border:1px solid var(--border);border-radius:6px;padding:5px 9px;font-size:.83rem;outline:none}}
.toolbar input{{width:180px}}
table{{width:100%;border-collapse:collapse;font-size:.83rem}}
th{{background:#f0f2f6;padding:9px 12px;text-align:left;font-weight:600;color:var(--muted);
    font-size:.75rem;text-transform:uppercase;letter-spacing:.04em;
    cursor:pointer;user-select:none;white-space:nowrap}}
th:hover{{background:#e5e8ef}}
th.sorted{{color:var(--text)}}
td{{padding:8px 12px;border-top:1px solid var(--border);vertical-align:middle}}
tr.data-row:hover td{{background:#f9fafb;cursor:pointer}}
.badge{{display:inline-block;padding:2px 7px;border-radius:999px;font-size:.72rem;font-weight:500;color:#fff}}
.num{{font-variant-numeric:tabular-nums;text-align:right}}
.muted-text{{color:var(--muted);font-style:italic;font-size:.8rem}}
.zero{{color:#ccc}}

/* Mini bar */
.mini-bar{{display:flex;height:8px;border-radius:4px;overflow:hidden;min-width:80px;background:#e9ecef}}
.mini-bar div{{flex-shrink:0}}

/* Detail panel */
tr.detail-row td{{padding:0;border-top:none}}
.detail-panel{{background:#fafbfc;border-top:2px solid #e0e4ea;padding:16px 20px;display:none}}
.detail-panel.open{{display:block}}
.cat-tabs{{display:flex;gap:0;border-bottom:2px solid var(--border);margin-bottom:14px}}
.cat-tab{{padding:7px 16px;cursor:pointer;font-size:.82rem;font-weight:600;border-bottom:3px solid transparent;margin-bottom:-2px;color:var(--muted)}}
.cat-tab.active{{border-bottom-color:var(--accent);color:var(--text)}}
.cat-tab:hover{{color:var(--text)}}
.cat-pane{{display:none}}.cat-pane.active{{display:block}}
.detail-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(260px,1fr));gap:14px}}
.detail-section{{background:#fff;border:1px solid var(--border);border-radius:8px;padding:12px 14px}}
.detail-section h4{{font-size:.78rem;font-weight:600;text-transform:uppercase;letter-spacing:.05em;color:var(--muted);margin-bottom:8px;display:flex;align-items:center;gap:6px}}
.detail-section h4 .cnt{{background:#e9ecef;color:var(--text);border-radius:999px;padding:1px 7px;font-size:.72rem}}
.file-list{{font-family:monospace;font-size:.74rem;max-height:220px;overflow-y:auto;line-height:1.75}}
.cap-note{{color:var(--muted);font-size:.72rem;margin-top:4px;font-style:italic}}
.info-row{{display:flex;flex-wrap:wrap;gap:12px;margin-bottom:10px;font-size:.8rem}}
.info-row span{{color:var(--muted)}}.info-row strong{{color:var(--text)}}
.ext-table{{font-size:.78rem;border-collapse:collapse;width:100%;margin-bottom:10px}}
.ext-table td{{padding:2px 6px}}.ext-table td:last-child{{text-align:right;font-variant-numeric:tabular-nums;color:var(--muted)}}
.ext-tag{{display:inline-block;background:#f0f2f6;border-radius:3px;padding:0px 5px;font-family:monospace;font-size:.73rem}}
.shared-badge{{display:inline-flex;align-items:center;gap:4px;background:#fff3cd;
               border:1px solid #ffc107;border-radius:4px;padding:2px 8px;font-size:.75rem;margin-bottom:8px}}
.type-filter-row{{display:flex;gap:6px;flex-wrap:wrap;margin-bottom:10px}}
.type-btn{{border:1px solid var(--border);border-radius:4px;padding:3px 9px;font-size:.75rem;cursor:pointer;background:#fff}}
.type-btn.active{{background:var(--accent);color:#fff;border-color:var(--accent)}}
</style>
</head>
<body>

<header>
  <h1>Yocto Source Report</h1>
  <p>Image: <strong>{image}</strong> &nbsp;&middot;&nbsp;
     Machine: <strong>{machine}</strong> &nbsp;&middot;&nbsp;
     Generated: {generated}</p>
</header>

<div class="wrap">

<!-- Summary cards -->
<div class="cards">
  <div class="card c4"><div class="val">{total_pkgs}</div><div class="lbl">Installed packages</div></div>
  <div class="card c3"><div class="val">{total_cu}</div><div class="lbl">Source files</div></div>
  <div class="card" style="border-left:3px solid #2980b9"><div class="val" style="color:#2980b9">{total_installed}</div><div class="lbl">Installed files</div></div>
  <div class="card" style="border-left:3px solid #e74c3c"><div class="val" style="color:#e74c3c">{total_copyleft}</div><div class="lbl">Copyleft packages</div></div>
  <div class="card" style="border-left:3px solid #8e44ad"><div class="val" style="color:#8e44ad">{total_patches}</div><div class="lbl">Patches</div></div>
  <div class="card" style="border-left:3px solid {sanity_color}"><div class="val" style="color:{sanity_color}">{sanity_issues}</div><div class="lbl">Sanity issues</div></div>
</div>

<!-- Charts -->
<div class="charts">
  <div class="cc">
    <h2>Source files per package</h2>
    <canvas id="barChart" height="300"></canvas>
  </div>
  <div class="cc">
    <h2>Package types</h2>
    <canvas id="donutChart"></canvas>
  </div>
</div>

<!-- Table -->
<div class="tbl-wrap">
  <div class="toolbar">
    <h2>All packages <span id="filteredCount" style="font-weight:400;color:#9ca3af"></span></h2>
    <input id="search" type="search" placeholder="Filter packages…">
    <select id="typeFilter">
      <option value="">All types</option>
      <option value="Userspace">Userspace</option>
      <option value="Kernel image">Kernel image</option>
      <option value="Kernel module">Kernel module</option>
      <option value="No compiled source">No compiled source</option>
    </select>
  </div>
  <table id="pkgTable">
    <thead><tr>
      <th data-col="name">Package <span class="si">↕</span></th>
      <th data-col="recipe">Recipe <span class="si">↕</span></th>
      <th data-col="type_label">Type <span class="si">↕</span></th>
      <th data-col="license">License <span class="si">↕</span></th>
      <th data-col="cu_own_srcs" style="text-align:right">Sources <span class="si">↕</span></th>
      <th data-col="installed_total" style="text-align:right">Installed <span class="si">↕</span></th>
      <th data-col="sanity_sort" style="text-align:center">Sanity <span class="si">↕</span></th>
    </tr></thead>
    <tbody id="tblBody"></tbody>
  </table>
</div>

</div>

<script>
const DATA = {data_json};
DATA.forEach(d=>{{d.sanity_sort=d.sanity?d.sanity.status==='FAIL'?2:d.sanity.status==='WARN'?1:0:0}});
// ── Charts ───────────────────────────────────────────────────────────────────
const barData = DATA.filter(d=>d.cu_own_srcs>0)
                    .sort((a,b)=>b.cu_own_srcs-a.cu_own_srcs);
new Chart(document.getElementById('barChart'),{{
  type:'bar',
  data:{{
    labels:barData.map(d=>d.name),
    datasets:[
      {{label:'Compiled Sources',data:barData.map(d=>d.cu_own_srcs),  backgroundColor:'#27ae60', stack:'s'}},
    ]
  }},
  options:{{
    indexAxis:'y', responsive:true,
    plugins:{{legend:{{position:'top'}},tooltip:{{mode:'index'}}}},
    scales:{{
      x:{{stacked:true,grid:{{color:'#f0f2f6'}}}},
      y:{{stacked:true,ticks:{{font:{{size:10}}}}}}
    }}
  }}
}});

const typeCounts={{}};
DATA.forEach(d=>{{typeCounts[d.type_label]=(typeCounts[d.type_label]||0)+1}});
new Chart(document.getElementById('donutChart'),{{
  type:'doughnut',
  data:{{
    labels:Object.keys(typeCounts),
    datasets:[{{data:Object.values(typeCounts),
      backgroundColor:['#4f86c6','#e07b39','#6dbf67','#b0b0b0'],
      borderWidth:2,borderColor:'#fff'}}]
  }},
  options:{{cutout:'62%',plugins:{{legend:{{position:'bottom',labels:{{font:{{size:11}},padding:12}}}}}}}}
}});

// ── Table ────────────────────────────────────────────────────────────────────
let sortCol='cu_own_srcs', sortAsc=false, filterText='', filterType='';

function fmt(n){{
  return n===0?'<span class="zero">0</span>':n.toLocaleString();
}}

function sanityBadge(s){{
  if(!s||s.status==='OK') return '<span style="color:#27ae60;font-weight:700">\u2713</span>';
  if(s.status==='FAIL') return '<span class="badge" style="background:#e74c3c">\u2717</span>';
  return '<span class="badge" style="background:#e67e22">\u26A0</span>';
}}

function extTable(ext){{
  if(!ext||!Object.keys(ext).length) return '<em style="color:#ccc;font-size:.75rem">—</em>';
  return '<table class="ext-table">' +
    Object.entries(ext).slice(0,12).map(([e,c])=>
      `<tr><td><span class="ext-tag">${{e}}</span></td><td>${{c}}</td></tr>`
    ).join('') +
    (Object.keys(ext).length>12?`<tr><td colspan="2" style="color:var(--muted);font-size:.72rem">… +${{Object.keys(ext).length-12}} more types</td></tr>`:'') +
    '</table>';
}}

function makeTypeFilter(files, panelId){{
  const exts = [...new Set(files.map(f=>{{
    const m=f.match(/\\.([^.]+)$/); return m?'.'+m[1]:'(none)';
  }}))] .sort();
  if(exts.length<=1) return '';
  const btns = exts.map(e=>`<button class="type-btn" data-ext="${{e}}" onclick="toggleTypeFilter(this,'${{panelId}}')">${{e}}</button>`).join('');
  return `<div class="type-filter-row" id="tfr-${{panelId}}">${{btns}}</div>`;
}}

window.toggleTypeFilter=function(btn,panelId){{
  btn.classList.toggle('active');
  const active=[...document.querySelectorAll(`#tfr-${{panelId}} .type-btn.active`)].map(b=>b.dataset.ext);
  const list=document.querySelector(`#fl-${{panelId}}`);
  if(!list) return;
  list.querySelectorAll('div[data-ext]').forEach(el=>{{
    el.style.display = (!active.length||active.includes(el.dataset.ext))?'':'none';
  }});
}};

function fileListFiltered(files, color, total, panelId){{
  if(!files||!files.length) return '<em style="color:#ccc;font-size:.75rem">none</em>';
  const items = files.map(f=>{{
    const m=f.match(/\\.([^.]+)$/); const ext=m?'.'+m[1]:'(none)';
    return `<div style="color:${{color}}" data-ext="${{ext}}">${{f}}</div>`;
  }}).join('');
  const cap = total>files.length
    ? `<div class="cap-note">Showing ${{files.length}} of ${{total}} files</div>`:'';
  return `<div class="file-list" id="fl-${{panelId}}">${{items}}</div>${{cap}}`;
}}

function pane(id, label, color, files, ext, total, active, srcBinMap){{
  const tfr = makeTypeFilter(files, id);
  // If srcBinMap provided, annotate files with their binary
  let fileItems;
  if(srcBinMap && Object.keys(srcBinMap).length){{
    fileItems = files.map(f=>{{
      const m=f.match(/\\.([^.]+)$/); const ext=m?'.'+m[1]:'(none)';
      const bins = srcBinMap[f];
      const annot = bins && bins.length
        ? ` <span style="color:#2980b9;font-size:.68rem;font-weight:500">← ${{bins.join(', ')}}</span>` : '';
      return `<div style="color:${{color}}" data-ext="${{ext}}">${{f}}${{annot}}</div>`;
    }}).join('');
    const cap = total>files.length
      ? `<div class="cap-note">Showing ${{files.length}} of ${{total}} files</div>`:'';
    fileItems = `<div class="file-list" id="fl-${{id}}">${{fileItems}}</div>${{cap}}`;
  }} else {{
    fileItems = fileListFiltered(files, color, total, id);
  }}
  return `<div class="cat-pane${{active?' active':''}}" id="cp-${{id}}">
    <div style="color:var(--muted);font-size:.78rem;margin-bottom:8px">
      <strong style="color:${{color}}">${{total.toLocaleString()}}</strong> files
    </div>
    ${{tfr}}
    <div class="detail-grid">
      <div class="detail-section">
        <h4>File types <span class="cnt">${{Object.keys(ext).length}}</span></h4>
        ${{extTable(ext)}}
      </div>
      <div class="detail-section" style="grid-column:span 2">
        <h4 style="color:${{color}}">${{label}} <span class="cnt">${{total}}</span></h4>
        ${{fileItems}}
      </div>
    </div>
  </div>`;
}}

function fmtSize(bytes){{
  if(bytes<1024) return bytes+' B';
  if(bytes<1024*1024) return (bytes/1024).toFixed(1)+' KB';
  return (bytes/1024/1024).toFixed(1)+' MB';
}}

function installedPane(id, files, total, elfCount, binarySources, active){{
  if(!files||!files.length) return `<div class="cat-pane${{active?' active':''}}" id="cp-${{id}}"><em style="color:#ccc">No installed files</em></div>`;
  const items = files.map(f=>{{
    const elfBadge = f.is_elf ? ' <span class="badge" style="background:#2980b9;font-size:.65rem">ELF</span>' : '';
    const srcs = binarySources[f.path];
    let srcList = '';
    if(srcs && srcs.length){{
      srcList = `<div style="margin-left:18px;color:#27ae60;font-size:.7rem">${{srcs.map(s=>'↳ '+s).join('<br>')}}</div>`;
    }}
    return `<div style="font-family:monospace;font-size:.74rem;line-height:1.75">
      <span style="color:var(--text)">${{f.path}}</span>
      <span style="color:var(--muted);font-size:.68rem;margin-left:6px">${{fmtSize(f.size)}}</span>
      ${{elfBadge}}${{srcList}}</div>`;
  }}).join('');
  const cap = total>files.length
    ? `<div class="cap-note">Showing ${{files.length}} of ${{total}} files</div>`:'';
  const mapped = Object.keys(binarySources).length;
  return `<div class="cat-pane${{active?' active':''}}" id="cp-${{id}}">
    <div style="color:var(--muted);font-size:.78rem;margin-bottom:8px">
      <strong style="color:#2980b9">${{total}}</strong> files · <strong>${{elfCount}}</strong> ELF binaries${{mapped?' · <strong>'+mapped+'</strong> with source mapping':''}}
    </div>
    <div class="detail-section" style="max-height:300px;overflow-y:auto">${{items}}</div>
    ${{cap}}
  </div>`;
}}

function licensePane(id, d, active){{
  if(!d.license && (!d.license_files||!d.license_files.length))
    return `<div class="cat-pane${{active?' active':''}}" id="cp-${{id}}"><em style="color:#ccc">No license information</em></div>`;
  const badge = d.copyleft ? '<span class="badge" style="background:#e74c3c;margin-left:6px">Copyleft</span>' : '';
  const oblColors = {{source_distribution:'#e74c3c',object_linking:'#e67e22',attribution:'#3498db',patent_grant:'#9b59b6',network_copyleft:'#c0392b',permissive:'#27ae60'}};
  const oblLabels = {{source_distribution:'Source Distribution',object_linking:'Object Linking',attribution:'Attribution',patent_grant:'Patent Grant',network_copyleft:'Network Copyleft',permissive:'Permissive'}};
  const obs = (d.obligations&&d.obligations.obligations)||[];
  const oblBadges = obs.map(o=>`<span class="badge" style="background:${{oblColors[o]||'#888'}};margin:2px 3px">${{oblLabels[o]||o}}</span>`).join('');
  const licFiles = (d.license_files||[]).map(f=>
    `<div style="font-family:monospace;font-size:.74rem;line-height:1.75">
      <span>${{f.name}}</span>
      <span style="color:var(--muted);font-size:.68rem;margin-left:6px">${{fmtSize(f.size)}}</span>
    </div>`
  ).join('');
  return `<div class="cat-pane${{active?' active':''}}" id="cp-${{id}}">
    <div style="margin-bottom:10px;font-size:.85rem">
      <strong>License:</strong> ${{d.license||'<em style="color:#ccc">unknown</em>'}}${{badge}}
    </div>
    ${{obs.length?`<div style="margin-bottom:10px">${{oblBadges}}</div>`:''}}
    <div class="detail-section">
      <h4>License files <span class="cnt">${{(d.license_files||[]).length}}</span></h4>
      ${{licFiles||'<em style="color:#ccc;font-size:.75rem">none collected</em>'}}
    </div>
  </div>`;
}}

function copyrightPane(id, d, active){{
  const crs = d.copyrights||[];
  if(!crs.length)
    return `<div class="cat-pane${{active?' active':''}}" id="cp-${{id}}"><em style="color:#ccc">No copyright notices extracted</em></div>`;
  const items = crs.map(c=>
    `<div style="font-size:.78rem;line-height:1.75">
      <span style="color:var(--muted);margin-right:8px">${{c.year}}</span>
      <span>${{c.holder}}</span>
    </div>`
  ).join('');
  return `<div class="cat-pane${{active?' active':''}}" id="cp-${{id}}">
    <div style="color:var(--muted);font-size:.78rem;margin-bottom:8px">
      <strong style="color:#2c3e50">${{crs.length}}</strong> copyright notices
    </div>
    <div class="detail-section" style="max-height:300px;overflow-y:auto">${{items}}</div>
  </div>`;
}}

function patchesPane(id, d, active){{
  const patches = d.patches||[];
  if(!patches.length)
    return `<div class="cat-pane${{active?' active':''}}" id="cp-${{id}}"><em style="color:#ccc">No patches</em></div>`;
  const items = patches.map(p=>
    `<div style="font-family:monospace;font-size:.74rem;line-height:1.75">
      <span style="color:#8e44ad">${{p.name}}</span>
      <span style="color:var(--muted);font-size:.68rem;margin-left:6px">${{fmtSize(p.size)}}</span>
    </div>`
  ).join('');
  return `<div class="cat-pane${{active?' active':''}}" id="cp-${{id}}">
    <div style="color:var(--muted);font-size:.78rem;margin-bottom:8px">
      <strong style="color:#8e44ad">${{patches.length}}</strong> patches
    </div>
    <div class="detail-section" style="max-height:300px;overflow-y:auto">${{items}}</div>
  </div>`;
}}

function metadataPane(id, d, active){{
  const rows = [
    ['Recipe', d.recipe],
    ['Version', d.version],
    ['License', (d.license||'') + (d.copyleft?' <span class="badge" style="background:#e74c3c">Copyleft</span>':'')],
    ['Homepage', d.homepage ? `<a href="${{d.homepage}}" target="_blank" style="color:var(--accent)">${{d.homepage}}</a>` : ''],
    ['Summary', d.summary||''],
    ['Description', d.description||''],
    ['Section', d.section||''],
    ['Runtime deps', d.rdepends||''],
  ].filter(r=>r[1]);
  const tbl = rows.map(r=>`<tr><td style="font-weight:600;color:var(--muted);white-space:nowrap;padding:4px 12px 4px 0;vertical-align:top">${{r[0]}}</td><td style="padding:4px 0">${{r[1]}}</td></tr>`).join('');
  return `<div class="cat-pane${{active?' active':''}}" id="cp-${{id}}">
    <table style="font-size:.82rem;border-collapse:collapse;width:100%">${{tbl}}</table>
  </div>`;
}}

function srcUriPane(id, d, active){{
  const files = d.src_uri_files||[];
  if(!files.length)
    return `<div class="cat-pane${{active?' active':''}}" id="cp-${{id}}"><em style="color:#ccc">No Yocto additional sources</em></div>`;
  const items = files.map(f=>
    `<div style="font-family:monospace;font-size:.74rem;line-height:1.75">
      <span style="color:#16a085">${{f.name}}</span>
      <span style="color:var(--muted);font-size:.68rem;margin-left:6px">${{fmtSize(f.size)}}</span>
    </div>`
  ).join('');
  return `<div class="cat-pane${{active?' active':''}}" id="cp-${{id}}">
    <div style="color:var(--muted);font-size:.78rem;margin-bottom:8px">
      <strong style="color:#16a085">${{files.length}}</strong> Yocto additional sources (scripts, configs, service files)
    </div>
    <div class="detail-section" style="max-height:300px;overflow-y:auto">${{items}}</div>
  </div>`;
}}

function depsPane(id, d, active){{
  const libs = d.needed_libs||[];
  if(!libs.length)
    return `<div class="cat-pane${{active?' active':''}}" id="cp-${{id}}"><em style="color:#ccc">No shared library dependencies detected</em></div>`;
  const hdr = `<tr><th style="text-align:left;padding:4px 8px;font-size:.73rem;color:var(--muted)">Library</th>
    <th style="text-align:left;padding:4px 8px;font-size:.73rem;color:var(--muted)">Provider</th>
    <th style="text-align:left;padding:4px 8px;font-size:.73rem;color:var(--muted)">License</th></tr>`;
  const trs = libs.map(l=>{{
    const cpBadge = l.copyleft ? ' <span class="badge" style="background:#e74c3c;font-size:.65rem">Copyleft</span>' : '';
    return `<tr>
      <td style="font-family:monospace;font-size:.74rem;padding:3px 8px">${{l.lib}}</td>
      <td style="font-size:.78rem;padding:3px 8px">${{l.provider_recipe||l.provider_pkg||'<span style="color:#ccc">unknown</span>'}}</td>
      <td style="font-size:.78rem;padding:3px 8px">${{l.provider_license||'<span style="color:#ccc">—</span>'}}${{cpBadge}}</td>
    </tr>`;
  }}).join('');
  const chain = d.linking_chain||{{}};
  const chainKeys = Object.keys(chain);
  const copyleftDeps = chainKeys.filter(k=>{{const u=chain[k].toUpperCase(); return ['GPL','LGPL','AGPL','EUPL','MPL','CDDL','OSL'].some(kw=>u.includes(kw));}});
  const chainNote = copyleftDeps.length
    ? `<div style="margin-top:10px;padding:8px 12px;background:#fff3cd;border:1px solid #ffc107;border-radius:6px;font-size:.78rem">
        <strong>⚠ Copyleft linking dependencies:</strong> ${{copyleftDeps.map(k=>`${{k}} (${{chain[k]}})`).join(', ')}}
      </div>` : '';
  return `<div class="cat-pane${{active?' active':''}}" id="cp-${{id}}">
    <div style="color:var(--muted);font-size:.78rem;margin-bottom:8px">
      <strong>${{libs.length}}</strong> shared libraries needed
    </div>
    <div class="detail-section" style="overflow-x:auto">
      <table style="border-collapse:collapse;width:100%">${{hdr}}${{trs}}</table>
    </div>
    ${{chainNote}}
  </div>`;
}}

function renderDetail(d, idx){{
  if(d.same_as)
    return `<div class="detail-panel open"><em>Sources shared with <strong>${{d.same_as}}</strong>.</em></div>`;

  const shared = d.shared_with&&d.shared_with.length
    ? `<div class="shared-badge">⚠ Shares recipe with: ${{d.shared_with.join(', ')}}</div>`:'';

  const info = `<div class="info-row">
    <span>Recipe: <strong>${{d.recipe}}</strong></span>
    <span>Version: <strong>${{d.version}}</strong></span>
    <span>Type: <strong>${{d.type_label}}</strong></span>
    ${{d.coverage?`<span>Compile log: <strong>${{d.coverage.covered}}/${{d.coverage.compile_total}}</strong> in binary</span>`:''}}
  </div>`;

  const cuPane   = pane(`cu-${{idx}}`,   'Source files',     '#27ae60', d.cu_files,  d.cu_ext,  d.cu_total,  !d.no_src, d.source_binaries);
  const instPane = installedPane(`inst-${{idx}}`, d.installed_files, d.installed_total, d.installed_elf, d.binary_sources||{{}}, d.no_src);
  const licPane  = licensePane(`lic-${{idx}}`, d, false);
  const crPane   = copyrightPane(`cr-${{idx}}`, d, false);
  const patPane  = patchesPane(`pat-${{idx}}`, d, false);
  const uriPane  = srcUriPane(`uri-${{idx}}`, d, false);
  const metPane  = metadataPane(`meta-${{idx}}`, d, false);
  const depPane  = depsPane(`dep-${{idx}}`, d, false);

  const nLibs = (d.needed_libs||[]).length;
  const nPat = (d.patches||[]).length;
  const nUri = (d.src_uri_files||[]).length;
  const nCr  = (d.copyrights||[]).length;

  const sanityBanner = d.sanity && d.sanity.issues && d.sanity.issues.length
    ? `<div style="padding:8px 12px;margin-bottom:10px;border-radius:6px;font-size:.8rem;border:1px solid ${{d.sanity.status==='FAIL'?'#e74c3c':'#e67e22'}};background:${{d.sanity.status==='FAIL'?'#fde8e8':'#fff3cd'}}">
        <strong>${{d.sanity.status==='FAIL'?'\u2717 Sanity FAIL':'\u26A0 Sanity warning'}}:</strong> ${{d.sanity.issues.join('; ')}}
      </div>` : '';

  return `<div class="detail-panel open">
    ${{sanityBanner}}${{info}}${{shared}}
    <div class="cat-tabs" data-idx="${{idx}}">
      <div class="cat-tab${{!d.no_src?' active':''}}" data-pane="cp-cu-${{idx}}" onclick="switchTab(this)">
        Sources <strong style="color:#27ae60">${{d.cu_own_srcs}}</strong> <span style="color:var(--muted);font-size:.7rem">(+${{d.cu_total - d.cu_own_srcs}} headers)</span></div>
      <div class="cat-tab${{d.no_src?' active':''}}" data-pane="cp-inst-${{idx}}" onclick="switchTab(this)">
        Installed Files <strong style="color:#2980b9">${{d.installed_total}}</strong></div>
      <div class="cat-tab" data-pane="cp-lic-${{idx}}" onclick="switchTab(this)">
        License ${{d.copyleft?'<span class="badge" style="background:#e74c3c;font-size:.6rem;vertical-align:middle">Copyleft</span>':''}}</div>
      <div class="cat-tab" data-pane="cp-cr-${{idx}}" onclick="switchTab(this)">
        Copyright ${{nCr?'<strong style="color:#2c3e50">'+nCr+'</strong>':''}}</div>
      <div class="cat-tab" data-pane="cp-pat-${{idx}}" onclick="switchTab(this)">
        Patches ${{nPat?'<strong style="color:#8e44ad">'+nPat+'</strong>':''}}</div>
      <div class="cat-tab" data-pane="cp-uri-${{idx}}" onclick="switchTab(this)">
        Yocto Additional Sources ${{nUri?'<strong style="color:#16a085">'+nUri+'</strong>':''}}</div>
      <div class="cat-tab" data-pane="cp-meta-${{idx}}" onclick="switchTab(this)">
        Metadata</div>
      <div class="cat-tab" data-pane="cp-dep-${{idx}}" onclick="switchTab(this)">
        Dependencies ${{nLibs?'<strong style="color:#e67e22">'+nLibs+'</strong>':''}}</div>
    </div>
    ${{cuPane}}${{instPane}}${{licPane}}${{crPane}}${{patPane}}${{uriPane}}${{metPane}}${{depPane}}
  </div>`;
}}

window.switchTab=function(tabEl){{
  const tabs = tabEl.closest('.cat-tabs');
  const panel = tabs.parentElement;
  tabs.querySelectorAll('.cat-tab').forEach(t=>t.classList.remove('active'));
  panel.querySelectorAll('.cat-pane').forEach(p=>p.classList.remove('active'));
  tabEl.classList.add('active');
  const target = document.getElementById(tabEl.dataset.pane);
  if(target) target.classList.add('active');
}};

function renderTable(){{
  const q=filterText.toLowerCase();
  let rows=DATA.filter(d=>
    (!q||d.name.includes(q)||d.recipe.includes(q))&&
    (!filterType||d.type_label===filterType)
  );
  rows.sort((a,b)=>{{
    let av=a[sortCol],bv=b[sortCol];
    if(typeof av==='string') av=av.toLowerCase(),bv=bv.toLowerCase();
    return sortAsc?(av>bv?1:-1):(av<bv?1:-1);
  }});

  document.getElementById('filteredCount').textContent=
    rows.length<DATA.length?`(${{rows.length}} of ${{DATA.length}})`:'';

  const tbody=document.getElementById('tblBody');
  tbody.innerHTML=rows.map((d,idx)=>{{
    const cpBadge = d.copyleft ? ' <span class="badge" style="background:#e74c3c;font-size:.6rem">Copyleft</span>' : '';
    const licText = d.license ? (d.license.length>35 ? d.license.substring(0,35)+'…' : d.license) : '<span style="color:#ccc">—</span>';
    return `<tr class="data-row" data-name="${{d.name}}">
      <td><strong>${{d.name}}</strong></td>
      <td style="color:#555">${{d.recipe}}</td>
      <td><span class="badge" style="background:${{d.color}}">${{d.type_label}}</span></td>
      <td style="font-size:.78rem">${{licText}}${{cpBadge}}</td>
      <td class="num" style="color:#27ae60">${{fmt(d.cu_own_srcs)}}</td>
      <td class="num" style="color:#2980b9">${{fmt(d.installed_total)}}</td>
      <td style="text-align:center">${{sanityBadge(d.sanity)}}</td>
    </tr>
    <tr class="detail-row"><td colspan="7">${{renderDetail(d,idx)}}</td></tr>`;
  }}).join('');

  tbody.querySelectorAll('tr.data-row').forEach(tr=>{{
    tr.addEventListener('click',()=>{{
      const panel=tr.nextElementSibling.querySelector('.detail-panel');
      if(!panel) return;
      const isOpen=panel.classList.contains('open');
      tbody.querySelectorAll('.detail-panel.open').forEach(p=>p.classList.remove('open'));
      if(!isOpen) panel.classList.add('open');
    }});
  }});

  document.querySelectorAll('th[data-col]').forEach(th=>{{
    th.classList.toggle('sorted',th.dataset.col===sortCol);
    const ic=th.querySelector('.si');
    if(ic) ic.textContent=th.dataset.col===sortCol?(sortAsc?'↑':'↓'):'↕';
  }});
}}

document.querySelectorAll('th[data-col]').forEach(th=>{{
  th.addEventListener('click',()=>{{
    sortCol===th.dataset.col?sortAsc=!sortAsc:(sortCol=th.dataset.col,sortAsc=false);
    renderTable();
  }});
}});
document.getElementById('search').addEventListener('input',e=>{{filterText=e.target.value;renderTable()}});
document.getElementById('typeFilter').addEventListener('change',e=>{{filterType=e.target.value;renderTable()}});

renderTable();
</script>
</body>
</html>
"""


# ═══════════════════════════════════════════════════════════════════════════════
# Reporter — generate an interactive HTML report of collected sources
# ═══════════════════════════════════════════════════════════════════════════════

def _list_files_in(d: Path) -> list[dict]:
    if not d.exists():
        return []
    result = []
    for f in d.rglob("*"):
        if not f.is_file():
            continue
        rel = str(f.relative_to(d))
        ext = f.suffix.lower() or "(no ext)"
        result.append({"path": rel, "ext": ext})
    return result


def _build_used_files_set(row: dict, pkg: "PackageInfo | None") -> "set[str] | None":
    """Build set of confirmed-used source file paths (relative to collected dir).

    Returns None when ALL collected files are confirmed (kernel_image only).
    Returns a set of confirmed paths for all other package types.

    Evidence sources:
      - DWARF CU paths: per-package (from this package's ELF binaries). Always safe.
      - debugsources.list: per-recipe. Only used for single-package recipes (no siblings).
      - compile log: per-recipe. Only used for single-package recipes (no siblings).
      - kernel_mod_obj_rels: per-module .o → .c/.S mapping from .mod files.

    For split-package recipes (shared work_ver), only per-package DWARF evidence
    is used — recipe-level sources would falsely confirm files that belong to
    sibling packages.

    For kernel modules, only the .c/.S compiled sources (from .mod → .o mapping)
    are confirmed. Headers from .cmd deps are shared kernel infrastructure and
    are deselected (they are already confirmed in kernel-image-image).
    """
    pkg_type = row["type"]

    # kernel_image: all files are the master set, all confirmed
    if pkg_type == "kernel_image":
        return None

    # Kernel modules: only .c/.S from .mod obj_rels are confirmed
    if pkg_type == "kernel_module":
        if not pkg or not pkg.kernel_mod_obj_rels:
            return set()  # No .mod evidence → nothing confirmed
        evidence: set[str] = set()
        for obj_rel in pkg.kernel_mod_obj_rels:
            p = Path(obj_rel)
            for ext in (".c", ".S"):
                evidence.add(str(p.parent / (p.stem + ext)))
        return evidence

    if pkg_type != "userspace" or not pkg:
        return set()  # Unknown type → nothing confirmed

    has_siblings = bool(row.get("shared_with"))
    evidence = set()

    # Stage 1: debugsources.list (per-recipe — only for single-package recipes)
    if not has_siblings and pkg.work_ver:
        dbgsrc = pkg.work_ver / "debugsources.list"
        if dbgsrc.exists():
            entries = read_debugsources(dbgsrc)
            prefix = f"/usr/src/debug/{pkg.recipe}/{pkg.ver}/"
            for e in entries:
                if e.startswith(prefix):
                    evidence.add(strip_src_root(e[len(prefix):]))

    # Stage 2: DWARF CU paths (per-package — always safe)
    dwarf_cu_rels = row.get("dwarf_cu_rels")
    if dwarf_cu_rels:
        evidence.update(dwarf_cu_rels)

    # Stage 3: compile log (per-recipe — only for single-package recipes)
    if not has_siblings and pkg.work_ver:
        log = pkg.work_ver / "temp" / "log.do_compile"
        if log.exists():
            cwd = _get_initial_cwd(pkg.work_ver)
            for cmd in parse_compile_log(log, cwd):
                try:
                    rel = str(cmd.src.relative_to(pkg.work_ver))
                    evidence.add(strip_src_root(rel))
                except ValueError:
                    pass

    if not evidence:
        return set()  # No evidence → nothing confirmed

    return evidence


def _ext_counts(files: list[dict]) -> dict[str, int]:
    c = Counter(f["ext"] for f in files)
    return dict(sorted(c.items(), key=lambda x: -x[1]))


def _dwarf_cross_check(installed_files: list[dict], pkg_split: Path,
                        cu_file_set: set[str],
                        recipe: str = "", ver: str = "") -> dict:
    binary_sources: dict[str, list[str]] = {}
    source_binaries: dict[str, list[str]] = {}
    dwarf_cu_rels: set[str] = set()

    # Pre-index collected files by basename for O(1) lookup
    cu_by_basename: dict[str, list[str]] = {}
    for p in cu_file_set:
        cu_by_basename.setdefault(os.path.basename(p), []).append(p)

    prefix = f"/usr/src/debug/{recipe}/{ver}/" if recipe else ""

    for f in installed_files:
        if not f["is_elf"]:
            continue
        elf_path = pkg_split / f["path"].lstrip("/")
        if not elf_path.exists():
            continue
        debug = find_debug_counterpart(elf_path, pkg_split)
        if not debug:
            continue
        dwarf_srcs = extract_dwarf_cu_sources(debug)
        matched: list[str] = []
        for dpath in dwarf_srcs:
            dbase = os.path.basename(dpath)
            matched.extend(cu_by_basename.get(dbase, []))
            # Collect raw DWARF CU relative paths for used-files evidence
            if prefix and dpath.startswith(prefix):
                dwarf_cu_rels.add(strip_src_root(dpath[len(prefix):]))
        matched = sorted(set(matched))
        if matched:
            binary_sources[f["path"]] = matched
            for src in matched:
                source_binaries.setdefault(src, []).append(f["path"])

    return {"binary_sources": binary_sources, "source_binaries": source_binaries,
            "dwarf_cu_rels": dwarf_cu_rels}


def _sanity_check(row: dict) -> dict:
    issues: list[str] = []
    pkg_type = row["type"]
    cu_own = row.get("cu_own_srcs", 0)
    elf_count = row.get("installed_elf", 0)

    if pkg_type == "userspace":
        if elf_count > 0 and cu_own == 0:
            issues.append(f"Has {elf_count} ELF binaries but 0 compiled sources")
    elif pkg_type == "kernel_module":
        if cu_own == 0:
            issues.append("Kernel module with no source files")
    elif pkg_type == "kernel_image":
        if cu_own == 0:
            issues.append("Kernel image with no source files")
    elif pkg_type == "no_source":
        if elf_count > 0:
            issues.append(f"Classified as no_source but has {elf_count} ELF binaries")

    if issues:
        has_fail = pkg_type in ("kernel_module", "kernel_image") and cu_own == 0
        status = "FAIL" if has_fail else "WARN"
    else:
        status = "OK"

    return {"status": status, "issues": issues}


_abs_path_cache: dict[str, Path] = {}  # work_ver key -> source dir


def _find_abs_source_path(rel_path: str, pkg: "PackageInfo") -> str:
    """Reconstruct absolute build-tree path from a collected relative path.

    Uses cached source directory per work_ver to avoid repeated filesystem
    operations.  The cache is populated on first call per work_ver using
    ``_find_src_subdir()`` (which consults _WORKDIR_INFRA exclusion set).
    """
    if pkg.pkg_type in ("kernel_image", "kernel_module") and pkg.kernel:
        return str(pkg.kernel.src_dir / rel_path)
    if pkg.work_ver:
        wv_key = str(pkg.work_ver)
        if wv_key not in _abs_path_cache:
            src = _find_src_subdir(pkg.work_ver, pkg.recipe, ver=pkg.ver)
            _abs_path_cache[wv_key] = src if src else pkg.work_ver
        base = _abs_path_cache[wv_key]
        return str(base / rel_path)
    return rel_path


# ═══════════════════════════════════════════════════════════════════════════════
# Auditor — comprehensive completeness audit
# ═══════════════════════════════════════════════════════════════════════════════

def _audit_dwarf_sources(pkg: PackageInfo, build_dir: Path,
                         timeout: int = 180) -> tuple[set[str], list[str]]:
    """Extract DWARF CU paths from all installed ELFs; report timed-out ELFs."""
    all_sources: set[str] = set()
    timed_out_elfs: list[str] = []
    if not pkg.work_ver:
        return all_sources, timed_out_elfs

    pkg_split = pkg.work_ver / "packages-split" / pkg.yocto_pkg
    elfs = find_installed_elfs(pkg_split) if pkg_split.exists() else []
    for elf in elfs:
        sources, did_timeout = _extract_dwarf_safe(elf, timeout)
        if did_timeout:
            timed_out_elfs.append(elf.name)
        all_sources.update(sources)
        dbg = find_debug_counterpart(elf, pkg_split)
        if dbg:
            sources2, did_timeout2 = _extract_dwarf_safe(dbg, timeout)
            if did_timeout2:
                timed_out_elfs.append(dbg.name)
            all_sources.update(sources2)

    return all_sources, timed_out_elfs


def _audit_log_sources(pkg: PackageInfo) -> set[str]:
    """Parse compile log and return stripped relative source paths."""
    if not pkg.work_ver:
        return set()
    log_file = pkg.work_ver / "temp" / "log.do_compile"
    if not log_file.exists():
        return set()
    try:
        cwd = _get_initial_cwd(pkg.work_ver)
        cmds = parse_compile_log(log_file, cwd)
        result: set[str] = set()
        for cmd in cmds:
            try:
                rel = str(cmd.src.relative_to(pkg.work_ver))
                result.add(strip_src_root(rel))
            except ValueError:
                pass
        return result
    except Exception:
        return set()


def _audit_debugsources(pkg: PackageInfo) -> set[str]:
    """Read debugsources.list and return stripped relative paths."""
    if not pkg.work_ver:
        return set()
    # Try the standard location: pkgdata or packages-split debugsources.list
    dbgsrc_file = pkg.work_ver / "debugsources.list"
    if not dbgsrc_file.exists():
        # Also check packages-split/<recipe>-src/
        for candidate in [
            pkg.work_ver / "packages-split" / (pkg.recipe + "-src"),
            pkg.work_ver / "pkgdata" / "runtime" / (pkg.recipe + "-src"),
        ]:
            dsf = candidate / "debugsources.list" if candidate.is_dir() else candidate
            if dsf.exists():
                dbgsrc_file = dsf
                break
    if not dbgsrc_file.exists():
        return set()
    try:
        entries = read_debugsources(dbgsrc_file)
    except Exception:
        return set()
    prefix = f"/usr/src/debug/{pkg.recipe}/{pkg.ver}/"
    result: set[str] = set()
    for e in entries:
        if e.startswith(prefix):
            result.add(e[len(prefix):])
    return result


class Auditor:
    """Comprehensive completeness audit across all evidence sources."""

    def __init__(self, session: YoctoSession, sources_dir: Path | None = None,
                 json_out: Path | None = None, pkg_filter: set[str] | None = None,
                 dwarf_timeout: int = 180, fail_on_warn: bool = False):
        self.session = session
        self.sources_dir = sources_dir or session.sources_dir
        self.json_out = json_out or session.output_dir / "audit.json"
        self.pkg_filter = pkg_filter
        self.dwarf_timeout = dwarf_timeout
        self.fail_on_warn = fail_on_warn

    def run(self) -> int:
        packages = self.session.discover()
        if self.pkg_filter:
            packages = [p for p in packages
                        if p.installed_name in self.pkg_filter]

        results: list[AuditResult] = []
        for pkg in packages:
            dispatch = {
                "userspace":     self.audit_userspace,
                "kernel_image":  self.audit_kernel_image,
                "kernel_module": self.audit_kernel_module,
                "no_source":     self.audit_no_source,
            }
            handler = dispatch.get(pkg.pkg_type)
            if handler:
                results.append(handler(pkg))

        # Print results
        for r in results:
            ev = r.evidence
            ev_str = " ".join(f"{k}={v}" for k, v in sorted(ev.items()))
            if r.verdict == "PASS":
                print(f"  [{r.pkg_name}]  {r.pkg_type}  PASS  ({ev_str})")
            else:
                print(f"  [{r.pkg_name}]  {r.pkg_type}  {r.verdict}")
                for f in r.findings:
                    print(f"    {f.severity}  {f.code}: {f.detail}")

        # Summary
        n_pass = sum(1 for r in results if r.verdict == "PASS")
        n_warn = sum(1 for r in results if r.verdict == "WARN")
        n_fail = sum(1 for r in results if r.verdict == "FAIL")
        print(f"\nAUDIT SUMMARY: {n_pass} PASS / {n_warn} WARN / {n_fail} FAIL")
        if n_fail:
            fail_pkgs = [r for r in results if r.verdict == "FAIL"]
            reasons = ", ".join(
                f"{r.pkg_name} ({r.findings[0].code})" if r.findings
                else r.pkg_name
                for r in fail_pkgs
            )
            print(f"FAIL: {reasons}")

        # Write JSON
        self.json_out.parent.mkdir(parents=True, exist_ok=True)
        json_data = []
        for r in results:
            json_data.append({
                "package": r.pkg_name,
                "type": r.pkg_type,
                "verdict": r.verdict,
                "findings": [
                    {"severity": f.severity, "code": f.code, "detail": f.detail}
                    for f in r.findings
                ],
                "evidence": r.evidence,
            })
        self.json_out.write_text(json.dumps(json_data, indent=2) + "\n")
        print(f"\nJSON report: {self.json_out}")

        if n_fail > 0:
            return 1
        if self.fail_on_warn and n_warn > 0:
            return 1
        return 0

    def _compute_verdict(self, findings: list[AuditFinding]) -> str:
        if any(f.severity == "FAIL" for f in findings):
            return "FAIL"
        if any(f.severity == "WARN" for f in findings):
            return "WARN"
        return "PASS"

    def _check_collection(self, pkg: PackageInfo, expected: set[str],
                          findings: list[AuditFinding]) -> None:
        """Check if expected sources are present in collected output."""
        if not self.sources_dir.exists():
            return
        collected = list_collected_files(self.sources_dir, pkg.installed_name)
        if not collected and not expected:
            return
        collected_basenames = {os.path.basename(c) for c in collected}
        missing = set()
        for src in expected:
            basename = os.path.basename(src)
            ext = os.path.splitext(basename)[1]
            if ext not in _COMPILED_EXTS:
                continue
            if basename not in collected_basenames:
                missing.add(src)
        if missing:
            sample = sorted(missing)[:5]
            findings.append(AuditFinding(
                "FAIL", "MISSING_FROM_COLLECTION",
                f"{len(missing)} source(s) in evidence but not collected: "
                + ", ".join(sample)
                + (f" … and {len(missing) - 5} more" if len(missing) > 5 else "")
            ))

    def audit_userspace(self, pkg: PackageInfo) -> AuditResult:
        findings: list[AuditFinding] = []

        # Gather evidence
        dwarf_srcs, timed_out = _audit_dwarf_sources(
            pkg, self.session.build_dir, self.dwarf_timeout)
        log_srcs = _audit_log_sources(pkg)
        dbgsrc = _audit_debugsources(pkg)

        # Filter DWARF to same-recipe sources only (collector behaviour).
        # DWARF paths are absolute; own-recipe sources live under the recipe
        # prefix /usr/src/debug/<recipe>/<ver>/ or under work_ver.
        own_prefix = f"/usr/src/debug/{pkg.recipe}/{pkg.ver}/"
        dwarf_all_compiled = {s for s in dwarf_srcs
                              if os.path.splitext(s)[1] in _COMPILED_EXTS}
        dwarf_own = set()
        dwarf_foreign = set()
        for s in dwarf_all_compiled:
            if s.startswith(own_prefix):
                dwarf_own.add(s[len(own_prefix):])
            elif pkg.work_ver and s.startswith(str(pkg.work_ver)):
                try:
                    rel = str(Path(s).relative_to(pkg.work_ver))
                    dwarf_own.add(strip_src_root(rel))
                except ValueError:
                    dwarf_foreign.add(s)
            else:
                dwarf_foreign.add(s)

        log_compiled = {s for s in log_srcs
                        if os.path.splitext(s)[1] in _COMPILED_EXTS}

        # Count installed ELFs
        elf_count = 0
        if pkg.work_ver:
            pkg_split = pkg.work_ver / "packages-split" / pkg.yocto_pkg
            if pkg_split.exists():
                elf_count = len(find_installed_elfs(pkg_split))

        # DWARF timeout
        for elf_name in timed_out:
            findings.append(AuditFinding(
                "FAIL", "DWARF_TIMEOUT",
                f"readelf timed out on {elf_name}"))

        # No evidence at all
        if elf_count > 0 and not dwarf_own and not log_compiled:
            findings.append(AuditFinding(
                "FAIL", "NO_EVIDENCE",
                f"{elf_count} ELF(s) installed but both DWARF and compile log "
                "produced no source paths"))

        # In log but not DWARF (compare stripped basenames).
        # The compile log is per-RECIPE, so log-only sources often belong to
        # sibling packages or test binaries — informational, not actionable.
        log_only: set[str] = set()
        if log_compiled and dwarf_own:
            dwarf_basenames = {os.path.basename(d) for d in dwarf_own}
            log_only = {s for s in log_compiled
                        if os.path.basename(s) not in dwarf_basenames}
            if log_only:
                findings.append(AuditFinding(
                    "INFO", "IN_LOG_NOT_DWARF",
                    f"{len(log_only)} source(s) in compile log but "
                    "not in DWARF (sibling packages / tests / strip / LTO)"))

        # Check collection — only DWARF-confirmed own-recipe sources.
        # Log-only sources are excluded: they likely belong to sibling
        # sub-packages or test binaries not installed in this package.
        self._check_collection(pkg, dwarf_own, findings)

        evidence = {
            "DWARF": len(dwarf_own),
            "log": len(log_compiled),
            "dbgsrc": len(dbgsrc),
            "foreign": len(dwarf_foreign),
        }
        verdict = self._compute_verdict(findings)
        return AuditResult(pkg.installed_name, pkg.pkg_type, verdict,
                           findings, evidence)

    def audit_kernel_image(self, pkg: PackageInfo) -> AuditResult:
        findings: list[AuditFinding] = []
        obj_count = 0
        missing_src = 0

        if pkg.kernel:
            kbuild = pkg.kernel.build_dir
            ksrc = pkg.kernel.src_dir
            module_objs = pkg.kernel.module_objs()

            # Enumerate non-module .o files
            for o in kbuild.rglob("*.o"):
                if o in module_objs:
                    continue
                obj_count += 1
                stem = o.stem
                found = False
                for ext in (".c", ".S", ".s"):
                    # Check source dir relative to build dir
                    try:
                        rel = o.relative_to(kbuild)
                    except ValueError:
                        continue
                    src_candidate = ksrc / rel.parent / (stem + ext)
                    if src_candidate.exists():
                        found = True
                        break
                if not found:
                    missing_src += 1

            # Parse kernel compile log if available
            log_srcs: set[str] = set()
            if pkg.work_ver:
                log_file = pkg.work_ver / "temp" / "log.do_compile"
                if log_file.exists():
                    try:
                        cwd = _get_initial_cwd(pkg.work_ver)
                        cmds = parse_compile_log(log_file, cwd)
                        log_srcs = {str(cmd.src) for cmd in cmds}
                    except Exception:
                        pass

            if missing_src > 0:
                findings.append(AuditFinding(
                    "WARN", "KERNEL_OBJ_NO_SOURCE",
                    f"{missing_src}/{obj_count} .o files have no matching "
                    "source in kernel-source/"))

            # Check collection
            if self.sources_dir.exists():
                collected = list_collected_files(
                    self.sources_dir, pkg.installed_name)
                if obj_count > 0 and not collected:
                    findings.append(AuditFinding(
                        "FAIL", "MISSING_FROM_COLLECTION",
                        f"{obj_count} .o files but no sources collected"))
        else:
            findings.append(AuditFinding(
                "WARN", "NO_KERNEL_INFO",
                "KernelInfo not available; cannot cross-check"))
            log_srcs = set()

        evidence = {"obj": obj_count, "missing_src": missing_src,
                    "log": len(log_srcs)}
        verdict = self._compute_verdict(findings)
        return AuditResult(pkg.installed_name, pkg.pkg_type, verdict,
                           findings, evidence)

    def audit_kernel_module(self, pkg: PackageInfo) -> AuditResult:
        findings: list[AuditFinding] = []
        n_objs = len(pkg.kernel_mod_obj_rels)
        missing = 0
        oot_candidates = 0

        if n_objs == 0:
            findings.append(AuditFinding(
                "WARN", "NO_MOD_OBJECTS",
                "No .mod object relations found"))
        elif pkg.kernel:
            ksrc = pkg.kernel.src_dir
            kbuild = pkg.kernel.build_dir
            for obj_rel in pkg.kernel_mod_obj_rels:
                stem = Path(obj_rel).stem
                found = False
                for ext in (".c", ".S", ".s"):
                    for base in (ksrc, kbuild):
                        candidate = base / Path(obj_rel).parent / (stem + ext)
                        if candidate.exists():
                            found = True
                            break
                    if found:
                        break
                if not found:
                    missing += 1
                    # Out-of-tree check
                    if pkg.work_ver:
                        for ext in (".c", ".S", ".s"):
                            if list(pkg.work_ver.rglob(stem + ext)):
                                oot_candidates += 1
                                break

            if missing > 0 and oot_candidates > 0:
                findings.append(AuditFinding(
                    "FAIL", "OOT_MODULE_NO_SOURCE",
                    f"{oot_candidates} out-of-tree module source(s) found in "
                    "work dir but not in kernel-source/"))
            elif missing > 0:
                findings.append(AuditFinding(
                    "WARN", "MODULE_OBJ_NO_SOURCE",
                    f"{missing}/{n_objs} module .o files have no matching source"))

        # Check collection
        if self.sources_dir.exists():
            collected = list_collected_files(
                self.sources_dir, pkg.installed_name)
            if n_objs > 0 and not collected:
                findings.append(AuditFinding(
                    "FAIL", "MISSING_FROM_COLLECTION",
                    f"{n_objs} module objects but no sources collected"))

        evidence = {"mod_objs": n_objs, "missing": missing,
                    "oot": oot_candidates}
        verdict = self._compute_verdict(findings)
        return AuditResult(pkg.installed_name, pkg.pkg_type, verdict,
                           findings, evidence)

    _DATA_SPLIT_SUFFIXES = (
        "-locale-", "-locale", "-meta", "-conf", "-doc", "-dev",
        "-staticdev", "-src", "-dbg", "-ptest", "-bash-completion",
    )

    def audit_no_source(self, pkg: PackageInfo) -> AuditResult:
        findings: list[AuditFinding] = []

        if not pkg.work_ver:
            findings.append(AuditFinding(
                "WARN", "SSTATE_NO_WORKDIR",
                "work_ver missing (sstate?), cannot verify classification"))
            verdict = self._compute_verdict(findings)
            return AuditResult(pkg.installed_name, pkg.pkg_type, verdict,
                               findings, {})

        # Check for ELFs in this package's split directory
        pkg_split = pkg.work_ver / "packages-split" / pkg.yocto_pkg
        has_local_elfs = False
        if pkg_split.exists():
            elfs = find_installed_elfs(pkg_split)
            has_local_elfs = bool(elfs)
            if elfs:
                findings.append(AuditFinding(
                    "FAIL", "NO_SOURCE_HAS_ELFS",
                    f"Classified as no_source but has {len(elfs)} installed "
                    "ELF binary/ies"))

        # Determine if this is a data-only split package
        is_data_only_split = False
        if not has_local_elfs:
            if any(s in pkg.installed_name for s in self._DATA_SPLIT_SUFFIXES):
                is_data_only_split = True
            elif pkg.installed_name != pkg.recipe:
                is_data_only_split = True

        # Check for actual compile log with real commands
        log_file = pkg.work_ver / "temp" / "log.do_compile"
        if log_file.exists():
            try:
                cwd = _get_initial_cwd(pkg.work_ver)
                cmds = parse_compile_log(log_file, cwd)
                if cmds:
                    severity = "INFO" if is_data_only_split else "FAIL"
                    findings.append(AuditFinding(
                        severity, "NO_SOURCE_HAS_COMPILE_LOG",
                        f"Classified as no_source but log.do_compile has "
                        f"{len(cmds)} compile command(s)"
                        + (" (data-only split — inherited from parent recipe)"
                           if is_data_only_split else "")))
            except Exception:
                pass

        # Check split-recipe parent presence
        if pkg.recipe != pkg.installed_name:
            # This is likely a split package; check parent is in manifest
            packages = self.session.discover()
            parent_recipes = {p.recipe for p in packages
                              if p.pkg_type != "no_source"}
            if pkg.recipe not in parent_recipes:
                findings.append(AuditFinding(
                    "WARN", "SPLIT_RECIPE_PARENT_ABSENT",
                    f"Split-recipe parent '{pkg.recipe}' not in manifest "
                    "as a compiled package"))

        evidence: dict = {}
        verdict = self._compute_verdict(findings)
        return AuditResult(pkg.installed_name, pkg.pkg_type, verdict,
                           findings, evidence)


class Reporter:
    def __init__(self, session: YoctoSession):
        self.session = session
        self.out_dir = session.sources_dir
        self.output = session.output_dir / "report.html"

    def run(self) -> None:
        session = self.session

        if not self.out_dir.exists():
            raise SystemExit(f"Sources dir not found: {self.out_dir}\n"
                             "Run 'python3 yocto/source_audit.py collect' first.")

        print(f"Discovering packages…")
        packages = session.discover()

        print(f"Collecting data for {len(packages)} packages…")
        rows = self.collect_data(packages)

        html = self.render_html(rows)
        self.output.write_text(html)

        csv_path = self.generate_csv(rows, packages)

        xlsx_path = self.generate_xlsx(rows, packages)

        total_cu  = sum(r["cu_total"]  for r in rows)
        total_installed = sum(r["installed_total"] for r in rows)
        sanity_issues = sum(1 for r in rows if r.get("sanity", {}).get("status") in ("WARN", "FAIL"))

        print(f"Report: {self.output}")
        print(f"CSV:    {csv_path}")
        if xlsx_path:
            print(f"Excel:  {xlsx_path}")
        else:
            print("Excel:  (skipped — install openpyxl: pip install openpyxl)")
        print(f"  Source files      : {total_cu:,}")
        print(f"  Installed files   : {total_installed:,}")
        print(f"  Sanity issues     : {sanity_issues}")

    def collect_data(self, packages: list[PackageInfo]) -> list[dict]:
        build_dir = self.session.build_dir
        machine = self.session.machine
        tmpdir = self.session.tmpdir
        pkgdata_runtime = tmpdir / "pkgdata" / machine / "runtime"

        shlibs_map = build_shlibs_map(build_dir, machine, tmpdir)
        license_cache: dict[str, str] = {}
        licenses_dir = self.session.output_dir / "licenses"
        patches_dir = self.session.output_dir / "patches"
        srcuri_dir = self.session.output_dir / "src_uri"
        copyrights_dir = self.session.output_dir / "copyrights"

        work_ver_map: dict[str, list[str]] = {}
        for pkg in packages:
            if pkg.work_ver and pkg.pkg_type == "userspace":
                work_ver_map.setdefault(str(pkg.work_ver), []).append(pkg.installed_name)

        rows = []
        for pkg in sorted(packages, key=lambda p: p.installed_name):
            pkg_dir = self.out_dir / pkg.installed_name
            no_src  = pkg.pkg_type == "no_source"
            same_as = next(pkg_dir.glob("SAME_AS_*.txt"), None) if pkg_dir.exists() else None

            cu_files = _list_files_in(pkg_dir)
            cu_ext   = _ext_counts(cu_files)

            coverage: dict | None = None
            if pkg.pkg_type == "userspace":
                cov = check_coverage(pkg, self.out_dir)
                st  = cov.get("status", "")
                if st not in ("NO_WORK_DIR", "NO_LOG", "NO_CMDS"):
                    coverage = {
                        "compile_total": cov["total"],
                        "covered":       cov["covered"],
                    }

            shared_with: list[str] = []
            if pkg.work_ver and pkg.pkg_type == "userspace":
                peers = work_ver_map.get(str(pkg.work_ver), [])
                shared_with = [p for p in peers if p != pkg.installed_name]

            pkg_split = (pkg.work_ver / "packages-split" / pkg.yocto_pkg
                         if pkg.work_ver else None)
            installed_files = get_installed_files(
                pkgdata_runtime, pkg.yocto_pkg,
                pkg_split=pkg_split)

            xcheck: dict = {"binary_sources": {}, "source_binaries": {},
                           "dwarf_cu_rels": set()}
            elf_count = sum(1 for f in installed_files if f["is_elf"])
            if elf_count and pkg_split and cu_files:
                cu_path_set = {f["path"] for f in cu_files}
                xcheck = _dwarf_cross_check(installed_files, pkg_split,
                                            cu_path_set, pkg.recipe, pkg.ver)

            # License and metadata
            metadata = get_pkg_metadata(pkgdata_runtime, pkg.yocto_pkg)
            license_str = metadata.get("license", "")
            copyleft = is_copyleft(license_str) if license_str else False

            # License files from collected output
            lic_dir = licenses_dir / pkg.recipe
            license_file_list: list[dict] = []
            if lic_dir.exists():
                for f in sorted(lic_dir.iterdir()):
                    if f.is_file():
                        license_file_list.append({
                            "name": f.name,
                            "size": f.stat().st_size,
                        })

            # Patches from collected output
            pat_dir = patches_dir / pkg.recipe
            patch_list: list[dict] = []
            if pat_dir.exists():
                for f in sorted(pat_dir.iterdir()):
                    if f.is_file():
                        patch_list.append({
                            "name": f.name,
                            "size": f.stat().st_size,
                        })

            # SRC_URI files from collected output
            uri_dir = srcuri_dir / pkg.recipe
            srcuri_file_list: list[dict] = []
            if uri_dir.exists():
                for f in sorted(uri_dir.rglob("*")):
                    if f.is_file():
                        srcuri_file_list.append({
                            "name": str(f.relative_to(uri_dir)),
                            "size": f.stat().st_size,
                        })

            # Linking dependencies
            deps = resolve_linking_deps(
                installed_files, pkg_split, shlibs_map,
                pkgdata_runtime, license_cache)

            cu_own_srcs = sum(
                1 for f in cu_files
                if Path(f["path"]).suffix.lower() in _COMPILED_EXTS
            )

            # License obligations
            obligations = classify_license(license_str) if license_str else {
                "obligations": [], "has_source_distribution": False,
                "has_object_linking": False, "has_attribution": False,
                "has_patent_grant": False, "has_network_copyleft": False,
                "is_permissive_only": True, "licenses": [],
            }

            # Copyright notices
            cr_file = copyrights_dir / f"{pkg.recipe}.txt"
            copyrights: list[dict] = []
            if cr_file.exists():
                for line in cr_file.read_text().splitlines():
                    line = line.strip()
                    if not line:
                        continue
                    parts = line.split("  ", 1)
                    if len(parts) == 2:
                        copyrights.append({"year": parts[0], "holder": parts[1]})

            row = {
                "name":       pkg.installed_name,
                "recipe":     pkg.recipe,
                "version":    pkg.ver,
                "type":       pkg.pkg_type,
                "type_label": TYPE_LABEL.get(pkg.pkg_type, pkg.pkg_type),
                "color":      TYPE_COLOR.get(pkg.pkg_type, "#888"),
                "cu_total":   len(cu_files),
                "cu_own_srcs": cu_own_srcs,
                "cu_ext":     cu_ext,
                "cu_files":   [f["path"] for f in cu_files[:300]],
                "no_src":     no_src,
                "same_as":    same_as.stem.replace("SAME_AS_", "") if same_as else None,
                "coverage":   coverage,
                "shared_with": shared_with,
                "installed_files": installed_files[:500],
                "installed_total": len(installed_files),
                "installed_elf":   elf_count,
                "binary_sources":  xcheck["binary_sources"],
                "source_binaries": xcheck["source_binaries"],
                "dwarf_cu_rels":   sorted(xcheck["dwarf_cu_rels"]),
                "license":         license_str,
                "copyleft":        copyleft,
                "obligations":     obligations,
                "license_files":   license_file_list,
                "patches":         patch_list,
                "homepage":        metadata.get("homepage", ""),
                "description":     metadata.get("description", ""),
                "summary":         metadata.get("summary", ""),
                "section":         metadata.get("section", ""),
                "rdepends":        metadata.get("rdepends", ""),
                "src_uri_files":   srcuri_file_list,
                "copyrights":      copyrights,
                "needed_libs":     deps["needed_libs"],
                "linking_chain":   deps["linking_chain"],
            }
            row["sanity"] = _sanity_check(row)
            rows.append(row)
        return rows

    def render_html(self, rows: list[dict]) -> str:
        total_cu  = sum(r["cu_total"]  for r in rows)
        total_installed = sum(r["installed_total"] for r in rows)
        total_copyleft = sum(1 for r in rows if r.get("copyleft"))
        total_patches = sum(len(r.get("patches", [])) for r in rows)
        sanity_warn = sum(1 for r in rows if r.get("sanity", {}).get("status") == "WARN")
        sanity_fail = sum(1 for r in rows if r.get("sanity", {}).get("status") == "FAIL")
        sanity_issues = sanity_warn + sanity_fail

        image = (self.session.manifest_path.stem
                 if self.session.manifest_path
                 else "unknown")

        if sanity_fail > 0:
            sanity_color = "#e74c3c"
        elif sanity_warn > 0:
            sanity_color = "#e67e22"
        else:
            sanity_color = "#27ae60"

        return HTML_TEMPLATE.format(
            image      = image,
            machine    = self.session.machine,
            generated  = datetime.now().strftime("%Y-%m-%d %H:%M"),
            total_pkgs = len(rows),
            total_cu   = f"{total_cu:,}",
            total_installed = f"{total_installed:,}",
            total_copyleft = total_copyleft,
            total_patches = f"{total_patches:,}",
            sanity_issues = sanity_issues,
            sanity_color = sanity_color,
            data_json  = json.dumps(rows, indent=None),
        )

    def generate_csv(self, rows: list[dict], packages: list["PackageInfo"]) -> Path:
        pkg_map = {p.installed_name: p for p in packages}
        csv_path = self.session.output_dir / "report.csv"
        with open(csv_path, "w", newline="") as fh:
            writer = csv.writer(fh, delimiter=";")
            writer.writerow(["ID", "Recipe", "Version", "Package", "Deselected", "Sources"])
            row_id = 1
            for row in rows:
                name = row["name"]
                recipe = row["recipe"]
                version = row["version"]
                pkg = pkg_map.get(name)
                if row["type"] == "no_source":
                    writer.writerow([row_id, recipe, version, name,
                                     "NO_SOURCE_FILES_USED", ""])
                    row_id += 1
                    continue
                used = _build_used_files_set(row, pkg)
                pkg_dir = self.out_dir / name
                cu_files = _list_files_in(pkg_dir)
                for f in cu_files:
                    rel = f["path"]
                    if used is None or rel in used:
                        deselected = "False"
                    else:
                        deselected = "True"
                    abs_path = _find_abs_source_path(rel, pkg) if pkg else rel
                    writer.writerow([row_id, recipe, version, name,
                                     deselected, abs_path])
                    row_id += 1
                if not cu_files:
                    writer.writerow([row_id, recipe, version, name, "True", ""])
                    row_id += 1
        return csv_path

    def generate_xlsx(self, rows: list[dict],
                      packages: list["PackageInfo"]) -> Path | None:
        """Generate multi-sheet Excel workbook. Returns path or None."""
        if not _HAS_OPENPYXL:
            return None

        xlsx_path = self.session.output_dir / "report.xlsx"
        wb = openpyxl.Workbook()

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="1A1D23", end_color="1A1D23",
                                  fill_type="solid")
        header_align = Alignment(horizontal="left", vertical="center",
                                 wrap_text=True)
        thin_border = Border(
            bottom=Side(style="thin", color="E0E4EA"),
        )
        copyleft_fill = PatternFill(start_color="FDE8E8", end_color="FDE8E8",
                                    fill_type="solid")
        warn_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD",
                                fill_type="solid")
        fail_fill = PatternFill(start_color="FDE8E8", end_color="FDE8E8",
                                fill_type="solid")

        def _style_header(ws):
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

        def _auto_width(ws, max_width=50):
            for col_cells in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col_cells[0].column)
                for cell in col_cells[:100]:  # sample first 100 rows
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(
                    max(max_len + 2, 10), max_width)

        def _unique_recipe_rows(rows):
            """Yield one row per unique recipe (first occurrence)."""
            seen: set[str] = set()
            for r in rows:
                recipe = r["recipe"]
                if recipe not in seen:
                    seen.add(recipe)
                    yield r

        # ── Sheet 1: Summary ──
        ws = wb.active
        ws.title = "Summary"
        image = (self.session.manifest_path.stem
                 if self.session.manifest_path else "unknown")
        total_cu = sum(r["cu_total"] for r in rows)
        total_installed = sum(r["installed_total"] for r in rows)
        total_copyleft = sum(1 for r in rows if r.get("copyleft"))
        total_patches = sum(len(r.get("patches", [])) for r in rows)
        total_srcuri = sum(len(r.get("src_uri_files", [])) for r in rows)
        sanity_issues = sum(
            1 for r in rows
            if r.get("sanity", {}).get("status") in ("WARN", "FAIL"))

        summary_data = [
            ("Build Information", ""),
            ("Image", image),
            ("Machine", self.session.machine),
            ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("", ""),
            ("Summary", ""),
            ("Total packages", len(rows)),
            ("Source files", total_cu),
            ("Installed files", total_installed),
            ("Copyleft packages", total_copyleft),
            ("Patches", total_patches),
            ("Yocto additional sources", total_srcuri),
            ("Sanity issues", sanity_issues),
            ("", ""),
            ("Type Breakdown", "Count"),
        ]
        type_counts = Counter(r["type_label"] for r in rows)
        for tl, cnt in type_counts.most_common():
            summary_data.append((tl, cnt))

        summary_data.append(("", ""))
        summary_data.append(("Top Licenses", "Packages"))
        lic_counts = Counter(
            r["license"] for r in rows if r.get("license"))
        for lic, cnt in lic_counts.most_common(15):
            summary_data.append((lic, cnt))

        for label, val in summary_data:
            ws.append([label, val])

        # Bold section headers
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=1):
            cell = row[0]
            if cell.value in ("Build Information", "Summary",
                              "Type Breakdown", "Top Licenses"):
                cell.font = Font(bold=True, size=11)
        _auto_width(ws)

        # ── Sheet 2: Packages ──
        ws2 = wb.create_sheet("Packages")
        ws2.append(["Package", "Recipe", "Version", "Type", "License",
                     "Copyleft", "Obligations", "Sources", "Installed Files",
                     "Patches", "Yocto Additional Sources", "Sanity", "Homepage"])
        for r in rows:
            sanity_st = r.get("sanity", {}).get("status", "OK")
            obs = r.get("obligations", {})
            obs_str = ", ".join(obs.get("obligations", []))
            ws2.append([
                r["name"], r["recipe"], r["version"], r["type_label"],
                r.get("license", ""), "Yes" if r.get("copyleft") else "No",
                obs_str, r["cu_own_srcs"], r["installed_total"],
                len(r.get("patches", [])),
                len(r.get("src_uri_files", [])),
                sanity_st, r.get("homepage", ""),
            ])
        _style_header(ws2)
        # Conditional formatting
        for row_idx in range(2, ws2.max_row + 1):
            copyleft_cell = ws2.cell(row=row_idx, column=6)
            sanity_cell = ws2.cell(row=row_idx, column=12)
            if copyleft_cell.value == "Yes":
                copyleft_cell.fill = copyleft_fill
            if sanity_cell.value == "FAIL":
                sanity_cell.fill = fail_fill
            elif sanity_cell.value == "WARN":
                sanity_cell.fill = warn_fill
        _auto_width(ws2)

        # ── Sheet 3: Source Files ──
        # For kernel modules, emit only compiled sources (.c/.S) per recipe
        # (not all ~500 headers per module) to keep XLSX size manageable.
        ws3 = wb.create_sheet("Source Files")
        ws3.append(["Package", "Recipe", "Version", "File Path",
                     "Extension", "DWARF Confirmed"])
        pkg_map = {p.installed_name: p for p in packages}
        kernel_recipe_done: set[str] = set()
        for r in rows:
            name = r["name"]
            recipe = r["recipe"]
            version = r["version"]
            if r["type"] == "no_source":
                ws3.append([name, recipe, version, "",
                            "", "NO_SOURCE_FILES_USED"])
                continue
            if r["type"] == "kernel_module":
                # Emit one summary row per kernel module
                ws3.append([name, recipe, version,
                            f"({r['cu_total']} files — see CSV for full list)",
                            "", "Yes"])
                continue
            pkg = pkg_map.get(name)
            used = _build_used_files_set(r, pkg)
            pkg_dir = self.out_dir / name
            cu_files = _list_files_in(pkg_dir)
            for f in cu_files:
                rel = f["path"]
                ext = f["ext"]
                confirmed = "Yes" if (used is None or rel in used) else "No"
                ws3.append([name, recipe, version, rel, ext, confirmed])
            if not cu_files:
                ws3.append([name, recipe, version, "", "", "No"])
        _style_header(ws3)
        _auto_width(ws3)

        # ── Sheet 4: Licenses ──
        ws4 = wb.create_sheet("Licenses")
        _obl_flags = ["source_distribution", "object_linking", "attribution",
                       "patent_grant", "network_copyleft", "permissive"]
        ws4.append(["Recipe", "License", "Copyleft", "License Files"] +
                   [f.replace("_", " ").title() for f in _obl_flags])
        for r in _unique_recipe_rows(rows):
            recipe = r["recipe"]
            lic_files = "; ".join(
                f["name"] for f in r.get("license_files", []))
            obs = r.get("obligations", {})
            obs_set = set(obs.get("obligations", []))
            ws4.append([
                recipe, r.get("license", ""),
                "Yes" if r.get("copyleft") else "No", lic_files,
            ] + ["Yes" if f in obs_set else "No" for f in _obl_flags])
            if r.get("copyleft"):
                for cell in ws4[ws4.max_row]:
                    cell.fill = copyleft_fill
        _style_header(ws4)
        _auto_width(ws4)

        # ── Sheet 5: Patches ──
        ws5 = wb.create_sheet("Patches")
        ws5.append(["Recipe", "Patch File", "Size (bytes)"])
        for r in _unique_recipe_rows(rows):
            recipe = r["recipe"]
            for p in r.get("patches", []):
                ws5.append([recipe, p["name"], p["size"]])
        _style_header(ws5)
        _auto_width(ws5)

        # ── Sheet 6: Yocto Additional Sources ──
        ws6 = wb.create_sheet("Yocto Additional Sources")
        ws6.append(["Recipe", "File Name", "Size (bytes)"])
        for r in _unique_recipe_rows(rows):
            recipe = r["recipe"]
            for f in r.get("src_uri_files", []):
                ws6.append([recipe, f["name"], f["size"]])
        _style_header(ws6)
        _auto_width(ws6)

        # ── Sheet 7: Dependencies ──
        ws7 = wb.create_sheet("Dependencies")
        ws7.append(["Package", "Library", "Provider Package",
                     "Provider Recipe", "Provider License", "Copyleft"])
        for r in rows:
            for lib in r.get("needed_libs", []):
                ws7.append([
                    r["name"], lib["lib"],
                    lib.get("provider_pkg", ""),
                    lib.get("provider_recipe", ""),
                    lib.get("provider_license", ""),
                    "Yes" if lib.get("copyleft") else "No",
                ])
                if lib.get("copyleft"):
                    for cell in ws7[ws7.max_row]:
                        cell.fill = copyleft_fill
        _style_header(ws7)
        _auto_width(ws7)

        # ── Sheet 8: Copyrights ──
        ws8 = wb.create_sheet("Copyrights")
        ws8.append(["Recipe", "Year", "Copyright Holder"])
        for r in _unique_recipe_rows(rows):
            recipe = r["recipe"]
            for cr in r.get("copyrights", []):
                ws8.append([recipe, cr.get("year", ""), cr.get("holder", "")])
        _style_header(ws8)
        _auto_width(ws8)

        wb.save(str(xlsx_path))
        return xlsx_path



# ═══════════════════════════════════════════════════════════════════════════════
# Source Distribution Archiver
# ═══════════════════════════════════════════════════════════════════════════════

class Archiver:
    """Create per-recipe source distribution tarballs for GPL compliance."""

    def __init__(self, session: "YoctoSession", copyleft_only: bool = True):
        self.session = session
        self.copyleft_only = copyleft_only
        self.archives_dir = session.output_dir / "archives"

    def run(self) -> int:
        """Create archives. Returns count of archives created."""
        self.archives_dir.mkdir(parents=True, exist_ok=True)
        sources_dir = self.session.sources_dir
        licenses_dir = self.session.output_dir / "licenses"
        patches_dir = self.session.output_dir / "patches"
        srcuri_dir = self.session.output_dir / "src_uri"

        packages = self.session.discover()

        # Build recipe → pkg_names and recipe → license maps
        recipe_pkgs: dict[str, list[str]] = {}
        recipe_license: dict[str, str] = {}
        pkgdata_runtime = (self.session.tmpdir / "pkgdata"
                           / self.session.machine / "runtime")
        for pkg in packages:
            recipe_pkgs.setdefault(pkg.recipe, []).append(pkg.installed_name)
            if pkg.recipe not in recipe_license:
                meta = get_pkg_metadata(pkgdata_runtime, pkg.yocto_pkg)
                recipe_license[pkg.recipe] = meta.get("license", "")

        count = 0
        for recipe in sorted(recipe_pkgs):
            lic = recipe_license.get(recipe, "")
            if self.copyleft_only and not is_copyleft(lic):
                continue

            # Gather content directories
            content_dirs: list[tuple[str, Path]] = []
            for name in recipe_pkgs[recipe]:
                d = sources_dir / name
                if d.exists() and any(d.iterdir()):
                    content_dirs.append((f"sources/{name}", d))
            lic_d = licenses_dir / recipe
            if lic_d.exists() and any(lic_d.iterdir()):
                content_dirs.append(("licenses", lic_d))
            pat_d = patches_dir / recipe
            if pat_d.exists() and any(pat_d.iterdir()):
                content_dirs.append(("patches", pat_d))
            uri_d = srcuri_dir / recipe
            if uri_d.exists() and any(uri_d.iterdir()):
                content_dirs.append(("src_uri", uri_d))

            if not content_dirs:
                continue

            archive_path = self.archives_dir / f"{recipe}.tar.gz"
            with tarfile.open(str(archive_path), "w:gz") as tar:
                for arcname_prefix, dir_path in content_dirs:
                    for fpath in sorted(dir_path.rglob("*")):
                        if fpath.is_file():
                            arcname = f"{recipe}/{arcname_prefix}/{fpath.relative_to(dir_path)}"
                            tar.add(str(fpath), arcname=arcname)
            count += 1
            print(f"  {recipe}.tar.gz")

        return count


# ═══════════════════════════════════════════════════════════════════════════════
# CLI — unified entry point
# ═══════════════════════════════════════════════════════════════════════════════

def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="source-audit",
        description="Collect, verify, and report on Yocto image source files.",
    )
    sub = parser.add_subparsers(dest="command", required=True)

    p_collect = sub.add_parser("collect",
        help="Collect sources into ./output/sources/")
    add_common_args(p_collect)
    p_collect.add_argument("--clean", action="store_true",
        help="Remove output directory before collecting")

    p_verify = sub.add_parser("verify",
        help="DWARF cross-check + coverage check (+ optional compile/audit)")
    add_common_args(p_verify)
    p_verify.add_argument("--compile", action="store_true",
        help="Overlay collected sources and recompile each recipe via "
             "'bitbake <recipe> -c compile -f'. Proves collected files are "
             "valid compilable replacements. Requires sourced build environment.")
    p_verify.add_argument("-p", "--packages", metavar="PKG[,PKG…]", default="",
        help="Comma-separated list of installed package names to verify (default: all)")
    p_verify.add_argument("--audit", action="store_true",
        help="Run comprehensive completeness audit")
    p_verify.add_argument("--json", metavar="FILE", default=None,
        help="Write audit results JSON (default: output/audit.json; implies --audit)")
    p_verify.add_argument("--fail-on-warn", action="store_true",
        help="Exit 1 on WARN (not just FAIL; implies --audit)")
    p_verify.add_argument("--dwarf-timeout", type=int, default=180,
        metavar="N", help="Per-ELF DWARF timeout in seconds (default: 180; implies --audit)")

    p_report = sub.add_parser("report",
        help="Generate interactive HTML report")
    add_common_args(p_report)
    p_archive = sub.add_parser("archive",
        help="Create per-recipe source distribution tarballs")
    add_common_args(p_archive)
    p_archive.add_argument("--all-licenses", action="store_true",
        help="Archive all recipes, not just copyleft (default: copyleft only)")

    p_all = sub.add_parser("all",
        help="Collect, verify (with compile test), and generate report")
    add_common_args(p_all)
    p_all.add_argument("--clean", action="store_true",
        help="Remove output directory before collecting")
    p_all.add_argument("-p", "--packages", metavar="PKG[,PKG…]", default="",
        help="Comma-separated list of installed package names to verify (default: all)")

    return parser


def cmd_collect(args) -> int:
    session = YoctoSession.from_args(args)
    collector = Collector(session, clean=args.clean)
    failures = collector.run()
    return 1 if failures else 0


# ── Shared helpers for cmd_verify / cmd_all compile-test loops ──

def _resolve_src_base(
    session: "YoctoSession",
    rep: PackageInfo,
    recipe: str,
) -> Path | None:
    """Determine the source base directory for a recipe's compile test."""
    bb_s: Path | None = None
    renv = session.query_recipe_env(recipe)
    if renv.get("S"):
        p = Path(renv["S"])
        if p.is_dir():
            bb_s = p
    if rep.kernel and rep.kernel.src_dir:
        return rep.kernel.src_dir
    if rep.work_ver:
        return _find_src_subdir(rep.work_ver, recipe, ver=rep.ver, bitbake_s=bb_s)
    return bb_s


def _print_compile_result(result: dict) -> None:
    """Print the result of a bitbake compile test."""
    r_status = result["status"]
    if r_status == "PASS":
        n_repl = result.get("files_replaced", 0)
        print(f"  PASS (replaced {n_repl} files with collected copies)")
    elif r_status == "SKIP":
        print(f"  SKIP: {result.get('reason', '')}")
    else:
        print(f"  FAIL: {result.get('reason', '')}")
        stderr = result.get("stderr", "")
        if stderr:
            lines = stderr.strip().splitlines()
            tail = lines[-30:] if len(lines) > 30 else lines
            for line in tail:
                print(f"    {line}")


def cmd_verify(args) -> int:
    session = YoctoSession.from_args(args)
    session.print_header()

    pkg_filter = set(args.packages.split(",")) if args.packages else set()
    rc = 0

    # ── Phase 1: DWARF cross-check ──
    print("=== DWARF CROSS-CHECK ===\n")
    verifier = Verifier(session)
    verify_rc = verifier.run()
    if verify_rc:
        rc = 1

    # ── Phase 2: Coverage check ──
    print("\n=== COVERAGE CHECK ===\n")

    out_dir = session.sources_dir
    if pkg_filter:
        print(f"Filter    : {', '.join(sorted(pkg_filter))}")

    if not out_dir.exists():
        raise SystemExit(f"Sources directory not found: {out_dir}\n"
                         "Run 'python3 yocto/source_audit.py collect' first.")

    packages = session.discover()
    _warn_sstate(packages)
    testable = [p for p in packages
                if not pkg_filter or p.installed_name in pkg_filter]
    print(f"Discovered {len(packages)} packages, "
          f"{len(testable)} to test\n")
    print("=" * 72)

    # Check bitbake env early if --compile requested
    if args.compile:
        bb_err = _check_bitbake_env(session.build_dir)
        if bb_err:
            raise SystemExit(f"Cannot run compile test: {bb_err}")

    summary: dict[str, str] = {}

    # ── Pre-compute recipe→package-names map for recipe-level coverage ──
    recipe_all_names: dict[str, list[str]] = {}
    for pkg in testable:
        recipe_all_names.setdefault(pkg.recipe, []).append(pkg.installed_name)

    # Track which recipes already had their coverage checked (recipe-level)
    recipe_cov_cache: dict[str, dict] = {}

    for pkg in sorted(testable, key=lambda p: p.installed_name):
        print(f"\n[{pkg.installed_name}]  (recipe={pkg.recipe} type={pkg.pkg_type})")

        if pkg.pkg_type == "no_source":
            print("  SKIP: no compiled source")
            summary[pkg.installed_name] = "SKIP"
            continue

        if pkg.pkg_type == "userspace":
            # Use cached recipe-level coverage if available
            if pkg.recipe in recipe_cov_cache:
                cov = recipe_cov_cache[pkg.recipe]
            else:
                all_names = recipe_all_names.get(pkg.recipe, [])
                cov = check_coverage(pkg, out_dir, all_pkg_names=all_names)
                recipe_cov_cache[pkg.recipe] = cov
            status = cov.get("status", "ERROR")

            if status == "NO_WORK_DIR":
                print("  SKIP: work dir not found")
                summary[pkg.installed_name] = "SKIP"
                continue
            if status in ("NO_LOG", "NO_CMDS"):
                pkg_src_dir = out_dir / pkg.installed_name
                has_sources = (pkg_src_dir.exists()
                               and any(pkg_src_dir.iterdir()))
                if has_sources:
                    label = ("no log.do_compile" if status == "NO_LOG"
                             else "no compile commands in log")
                    print(f"  OK (sources collected, {label})")
                    summary[pkg.installed_name] = "OK"
                else:
                    label = ("no log.do_compile" if status == "NO_LOG"
                             else "no compile commands found in log")
                    print(f"  SKIP: {label}")
                    summary[pkg.installed_name] = "SKIP"
                continue

            total = cov["total"]
            covered = cov["covered"]
            not_coll = cov["not_collected"]
            outside = cov["outside"]

            print(f"  Compile commands : {total}")
            print(f"  Covered          : {covered}/{total}"
                  + ("  ✓" if covered == total else ""))
            if not_coll:
                print_list("Not collected (in work_ver)", not_coll)
            if outside:
                print_list("Outside work_ver (generated/external)", outside)

            summary[pkg.installed_name] = cov["status"]  # OK or INCOMPLETE

        elif pkg.pkg_type in ("kernel_image", "kernel_module"):
            pkg_src_dir = out_dir / pkg.installed_name
            if pkg_src_dir.is_dir():
                n_files = sum(1 for _ in pkg_src_dir.rglob("*") if _.is_file())
            else:
                n_files = 0
            if n_files == 0:
                print("  INCOMPLETE: 0 collected source files")
                summary[pkg.installed_name] = "INCOMPLETE"
            else:
                print(f"  Collected files: {n_files}")
                summary[pkg.installed_name] = "OK"

    # ── Phase 3: Bitbake compile test (when --compile) ──
    if args.compile:
        print("\n" + "=" * 72)
        print("BITBAKE COMPILE TEST")
        print("=" * 72)

        # Group packages by recipe (deduplicate)
        recipe_pkgs: dict[str, list[PackageInfo]] = {}
        for pkg in testable:
            if summary.get(pkg.installed_name) == "SKIP":
                continue
            recipe_pkgs.setdefault(pkg.recipe, []).append(pkg)

        for recipe, pkgs in sorted(recipe_pkgs.items()):
            pkg_names = [p.installed_name for p in pkgs]
            rep = pkgs[0]  # representative package

            src_base = _resolve_src_base(session, rep, recipe)
            if not src_base:
                print(f"\n[{recipe}]  SKIP: could not find source directory")
                for p in pkgs:
                    summary[p.installed_name] = "SKIP"
                continue

            print(f"\n[{recipe}]  packages={','.join(pkg_names)}")
            print(f"  src_base: {src_base}")

            result = bitbake_compile_test(
                recipe=recipe,
                src_base=src_base,
                sources_dir=out_dir,
                pkg_names=pkg_names,
                build_dir=session.build_dir,
                verbose=args.verbose,
                yocto_tmpdir=session.tmpdir,
                bb_env=session._bb_env,
            )

            _print_compile_result(result)

            # Propagate status to all packages from this recipe.
            # Compile PASS overrides INCOMPLETE — the bitbake compile
            # proves collected sources are valid; uncollected sources
            # are for non-installed binaries (tests, benchmarks, etc.).
            r_status = result["status"]
            for p in pkgs:
                if r_status == "FAIL":
                    summary[p.installed_name] = "FAIL"
                elif r_status == "PASS":
                    summary[p.installed_name] = "PASS"
                # SKIP leaves existing summary status unchanged

    print("\n" + "=" * 72)
    print("COVERAGE SUMMARY")
    print("=" * 72)

    ok_pkgs   = [p for p, s in summary.items() if s in ("OK", "PASS")]
    skip_pkgs = [p for p, s in summary.items() if s.startswith("SKIP")]
    warn_pkgs = [p for p, s in summary.items() if s == "INCOMPLETE"]
    fail_pkgs = [p for p, s in summary.items() if s == "FAIL"]

    print(f"  OK/PASS   : {len(ok_pkgs)}")
    print(f"  SKIP      : {len(skip_pkgs)}  (no source / no log)")
    print(f"  INCOMPLETE: {len(warn_pkgs)}  (some compiled sources not in collection)")
    for p in warn_pkgs:
        print(f"    {p}")
    print(f"  FAIL      : {len(fail_pkgs)}  (compile errors)")
    for p in fail_pkgs:
        print(f"    {p}")

    if fail_pkgs:
        rc = 1

    # ── Phase 4: Audit (when --audit or audit-implying flags) ──
    if args.audit or args.json or args.fail_on_warn or args.dwarf_timeout != 180:
        print("\n=== AUDIT ===\n")
        pkg_filter_set = (set(args.packages.split(","))
                          if args.packages else None)
        json_out = Path(args.json) if args.json else None
        auditor = Auditor(
            session,
            json_out=json_out,
            pkg_filter=pkg_filter_set,
            dwarf_timeout=args.dwarf_timeout,
            fail_on_warn=args.fail_on_warn,
        )
        audit_rc = auditor.run()
        if audit_rc:
            rc = 1

    return rc


def cmd_report(args) -> int:
    session = YoctoSession.from_args(args)
    reporter = Reporter(session)
    reporter.run()
    return 0


def cmd_archive(args) -> int:
    session = YoctoSession.from_args(args)
    copyleft_only = not getattr(args, "all_licenses", False)
    archiver = Archiver(session, copyleft_only=copyleft_only)
    print(f"\nCreating source archives ({'copyleft only' if copyleft_only else 'all licenses'})…\n")
    count = archiver.run()
    print(f"\nDone. {count} archives in {archiver.archives_dir}")
    return 0


def cmd_all(args) -> int:
    """Run collect → verify → add missing sources → compile test → report."""
    session = YoctoSession.from_args(args)

    bb_err = _check_bitbake_env(session.build_dir)
    if bb_err:
        raise SystemExit(f"Cannot run compile test: {bb_err}")

    pkg_filter = set(args.packages.split(",")) if args.packages else set()
    rc = 0

    # ── Stage 1: Collect ──
    print("\n" + "=" * 72)
    print("  STAGE 1: COLLECT")
    print("=" * 72 + "\n")
    collector = Collector(session, clean=args.clean)
    failures = collector.run()
    if failures:
        print(f"\n!! Collection failed for: {', '.join(name for name, _ in failures)}")
        return 1

    # ── Stage 2: DWARF cross-check ──
    print("\n" + "=" * 72)
    print("  STAGE 2: DWARF CROSS-CHECK")
    print("=" * 72 + "\n")
    session.print_header()
    verifier = Verifier(session)
    if verifier.run():
        rc = 1

    # ── Stage 3: Coverage check + add missing sources ──
    print("\n" + "=" * 72)
    print("  STAGE 3: COVERAGE CHECK + ADD MISSING SOURCES")
    print("=" * 72 + "\n")

    packages = session.discover()
    _warn_sstate(packages)
    testable = [p for p in packages
                if not pkg_filter or p.installed_name in pkg_filter]
    out_dir = session.sources_dir

    recipe_all_names: dict[str, list[str]] = {}
    for pkg in testable:
        recipe_all_names.setdefault(pkg.recipe, []).append(pkg.installed_name)

    recipe_cov_cache: dict[str, dict] = {}
    summary: dict[str, str] = {}
    total_added = 0

    for pkg in sorted(testable, key=lambda p: p.installed_name):
        print(f"\n[{pkg.installed_name}]  (recipe={pkg.recipe} type={pkg.pkg_type})")
        if pkg.pkg_type == "no_source":
            print("  SKIP: no compiled source")
            summary[pkg.installed_name] = "SKIP"
            continue
        if pkg.pkg_type in ("kernel_image", "kernel_module"):
            summary[pkg.installed_name] = "SKIP(kernel)"
            continue
        if pkg.pkg_type != "userspace":
            continue

        if pkg.recipe in recipe_cov_cache:
            cov = recipe_cov_cache[pkg.recipe]
        else:
            all_names = recipe_all_names.get(pkg.recipe, [])
            cov = check_coverage(pkg, out_dir, all_pkg_names=all_names)
            recipe_cov_cache[pkg.recipe] = cov

        status = cov.get("status", "?")
        summary[pkg.installed_name] = status

        if status in ("NO_WORK_DIR", "NO_LOG", "NO_CMDS"):
            print(f"  {status}")
            continue

        total = cov["total"]
        covered = cov["covered"]
        not_coll = cov["not_collected"]
        not_coll_abs = cov.get("not_collected_abs", {})
        print(f"  Covered: {covered}/{total}")

        # Copy missing source files into the collected directory
        if not_coll_abs:
            pkg_out = out_dir / pkg.installed_name
            pkg_out.mkdir(parents=True, exist_ok=True)
            added = 0
            for stripped, abs_path in not_coll_abs.items():
                dst = pkg_out / stripped
                if not dst.exists():
                    dst.parent.mkdir(parents=True, exist_ok=True)
                    shutil.copy2(abs_path, dst)
                    added += 1
            if added:
                print(f"  + Added {added} missing source files from work directory")
                total_added += added

    if total_added:
        print(f"\nTotal missing sources added: {total_added}")

    # ── Stage 4: Bitbake compile test ──
    print("\n" + "=" * 72)
    print("  STAGE 4: BITBAKE COMPILE TEST")
    print("=" * 72 + "\n")

    # Group packages by recipe (deduplicate); skip no_source + kernel
    recipe_pkgs: dict[str, list[PackageInfo]] = {}
    for pkg in testable:
        if summary.get(pkg.installed_name, "").startswith("SKIP"):
            continue
        recipe_pkgs.setdefault(pkg.recipe, []).append(pkg)

    for recipe, pkgs in sorted(recipe_pkgs.items()):
        pkg_names = [p.installed_name for p in pkgs]
        rep = pkgs[0]

        src_base = _resolve_src_base(session, rep, recipe)
        if not src_base:
            print(f"\n[{recipe}]  SKIP: could not find source directory")
            for p in pkgs:
                summary[p.installed_name] = "SKIP"
            continue

        print(f"\n[{recipe}]  packages={','.join(pkg_names)}")
        print(f"  src_base: {src_base}")

        result = bitbake_compile_test(
            recipe=recipe,
            src_base=src_base,
            sources_dir=out_dir,
            pkg_names=pkg_names,
            build_dir=session.build_dir,
            verbose=session.verbose,
            yocto_tmpdir=session.tmpdir,
            bb_env=session._bb_env,
        )

        _print_compile_result(result)

        r_status = result["status"]
        for p in pkgs:
            if r_status == "FAIL":
                summary[p.installed_name] = "FAIL"
            elif r_status == "PASS":
                summary[p.installed_name] = "PASS"

    # ── Compile summary ──
    print("\n" + "=" * 72)
    print("SUMMARY")
    print("=" * 72)
    ok_pkgs   = [p for p, s in summary.items() if s in ("OK", "PASS")]
    skip_pkgs = [p for p, s in summary.items() if s.startswith("SKIP")]
    warn_pkgs = [p for p, s in summary.items() if s == "INCOMPLETE"]
    fail_pkgs = [p for p, s in summary.items() if s == "FAIL"]
    print(f"  OK/PASS   : {len(ok_pkgs)}")
    print(f"  SKIP      : {len(skip_pkgs)}")
    print(f"  INCOMPLETE: {len(warn_pkgs)}")
    for p in warn_pkgs:
        print(f"    {p}")
    print(f"  FAIL      : {len(fail_pkgs)}")
    for p in fail_pkgs:
        print(f"    {p}")
    if fail_pkgs:
        rc = 1

    # ── Stage 5: Report ──
    print("\n" + "=" * 72)
    print("  STAGE 5: REPORT")
    print("=" * 72 + "\n")
    reporter = Reporter(session)
    reporter.run()

    # ── Stage 6: Archive ──
    print("\n" + "=" * 72)
    print("  STAGE 6: ARCHIVE (copyleft recipes)")
    print("=" * 72 + "\n")
    archiver = Archiver(session, copyleft_only=True)
    n_archives = archiver.run()
    print(f"\n  {n_archives} archives created in {archiver.archives_dir}")

    return rc


def main():
    parser = build_parser()
    args = parser.parse_args()

    handlers = {
        "collect": cmd_collect,
        "verify":  cmd_verify,
        "report":  cmd_report,
        "archive": cmd_archive,
        "all":     cmd_all,
    }

    handler = handlers.get(args.command)
    if handler is None:
        parser.print_help()
        sys.exit(1)

    sys.exit(handler(args))


if __name__ == "__main__":
    main()
